#!/usr/bin/env python3
"""
Marktplaats Product Manager - Registratie & voorraadbeheer

Vervangt de Google Form + Google Sheets-combinatie door een lokale app.
Deelt exact dezelfde kolomstructuur (A-X) als auto_marktplaats.py, zodat
beide apps dezelfde Google Sheet (of lokale XML) kunnen gebruiken.

Tab 1 "Registreren"  - nieuw product invoeren, artikelnummer + map +
                        omschrijving.txt + barcode genereren
Tab 2 "Overzicht"     - alle producten bekijken/bewerken/verwijderen,
                        markeren als verkocht, omzet-overzicht
"""

import gi
gi.require_version('Gtk', '3.0')
from gi.repository import Gtk, Gdk, GLib, GdkPixbuf
import os
import sys
import json
import hashlib
import secrets
import shutil
import subprocess
import datetime
import re
import csv
from io import StringIO

# ============================================
# KOLOMSTRUCTUUR (gedeeld met auto_marktplaats.py)
# ============================================
COLUMNS = [
    "artikelnummer",      # A (0)
    "titel",               # B (1)
    "categorie",            # C (2)
    "omschrijving",          # D (3)
    "online",                 # E (4)
    "lengte",                  # F (5)
    "breedte",                  # G (6)
    "hoogte",                     # H (7)
    "gewicht",                     # I (8)
    "conditie",                     # J (9)
    "staat_details",                 # K (10)
    "waarde_min",                     # L (11)
    "waarde_max",                      # M (12)
    "vraagprijs",                       # N (13)
    "aanmaakdatum",                      # O (14)
    "aanmaaktijd",                       # P (15)
    "tijdsperiode",                       # Q (16)
    "opslaglocatie",                       # R (17)
    "sublocatie",                           # S (18)
    "rij",                                   # T (19)
    "folder_locatie",                         # U (20)
    "verkocht",                                 # V (21)
    "verkoopprijs",                              # W (22)
    "verkoopdatum",                               # X (23)
    "algemene_voorwaarden",                        # Y (24)
    "url_1",                                       # Z (25)
    "url_2",                                       # AA (26)
    "url_3",                                       # AB (27)
    "url_4",                                       # AC (28)
    "url_5",                                       # AD (29)
    "leverwijze",                                   # AE (30)
    "klant_naam",                                   # AF (31)
    "klant_telefoon",                               # AG (32)
    "klant_email",                                   # AH (33)
    "ophaal_afspraak",                               # AI (34)
    "track_trace",                                   # AJ (35)
    "verwerkt_door",                                 # AK (36)
    "toegewezen_aan",                                # AL (37)
]
COL = {name: idx for idx, name in enumerate(COLUMNS)}

# ============================================
# KOLOM LABELS
# ============================================
KOLOM_LABELS = {
    "artikelnummer": "Artikelnummer",
    "titel": "Titel",
    "categorie": "Categorie",
    "omschrijving": "Omschrijving",
    "online": "Online",
    "lengte": "Lengte (cm)",
    "breedte": "Breedte (cm)",
    "hoogte": "Hoogte (cm)",
    "gewicht": "Gewicht (kg)",
    "conditie": "Conditie",
    "staat_details": "Staat details",
    "waarde_min": "Min. waarde (€)",
    "waarde_max": "Max. waarde (€)",
    "vraagprijs": "Vraagprijs (€)",
    "aanmaakdatum": "Aanmaakdatum",
    "aanmaaktijd": "Aanmaaktijd",
    "tijdsperiode": "Tijdsperiode",
    "opslaglocatie": "Opslaglocatie",
    "sublocatie": "Sublocatie",
    "rij": "Rij",
    "folder_locatie": "Folder locatie",
    "verkocht": "Verkocht",
    "verkoopprijs": "Verkoopprijs (€)",
    "verkoopdatum": "Verkoopdatum",
    "algemene_voorwaarden": "Algemene voorwaarden",
    "url_1": "URL 1",
    "url_2": "URL 2",
    "url_3": "URL 3",
    "url_4": "URL 4",
    "url_5": "URL 5",
    "leverwijze": "Leverwijze",
    "klant_naam": "Klant naam",
    "klant_telefoon": "Klant telefoon",
    "klant_email": "Klant e-mail",
    "ophaal_afspraak": "Ophaal afspraak",
    "track_trace": "Track & Trace",
    "verwerkt_door": "Verwerkt door",
    "toegewezen_aan": "Toegewezen aan",
}

# Standaard zichtbare kolommen per filter
DEFAULT_ZICHTBARE_KOLOMMEN_ALLES = [
    "artikelnummer", "titel", "categorie", "conditie", "vraagprijs",
    "aanmaakdatum", "aanmaaktijd", "verwerkt_door", "online", "verkocht",
    "verkoopprijs", "leverwijze", "klant_naam", "toegewezen_aan"
]

DEFAULT_ZICHTBARE_KOLOMMEN_OFFLINE = [
    "artikelnummer", "titel", "categorie", "conditie", "vraagprijs",
    "aanmaakdatum", "aanmaaktijd", "opslaglocatie", "sublocatie", "rij",
    "verwerkt_door", "toegewezen_aan"
]

DEFAULT_ZICHTBARE_KOLOMMEN_ONLINE = [
    "artikelnummer", "titel", "categorie", "conditie", "vraagprijs",
    "aanmaakdatum", "aanmaaktijd", "url_1", "url_2", "url_3",
    "verwerkt_door", "toegewezen_aan"
]

DEFAULT_ZICHTBARE_KOLOMMEN_VERKOCHT = [
    "artikelnummer", "titel", "categorie", "verkoopprijs", "verkoopdatum",
    "leverwijze", "klant_naam", "klant_telefoon", "klant_email",
    "ophaal_afspraak", "track_trace", "verwerkt_door", "toegewezen_aan"
]

# Filter types
FILTER_ALLES = "alles"
FILTER_OFFLINE = "offline"
FILTER_ONLINE = "online"
FILTER_VERKOCHT = "verkocht"

FILTER_NAMEN = {
    FILTER_ALLES: "Alles",
    FILTER_OFFLINE: "Offline",
    FILTER_ONLINE: "Online",
    FILTER_VERKOCHT: "Verkocht"
}

DEFAULT_ZICHTBARE_PER_FILTER = {
    FILTER_ALLES: DEFAULT_ZICHTBARE_KOLOMMEN_ALLES,
    FILTER_OFFLINE: DEFAULT_ZICHTBARE_KOLOMMEN_OFFLINE,
    FILTER_ONLINE: DEFAULT_ZICHTBARE_KOLOMMEN_ONLINE,
    FILTER_VERKOCHT: DEFAULT_ZICHTBARE_KOLOMMEN_VERKOCHT,
}

CONDITIES = ["Nieuwstaat", "Zo goed als nieuw", "Gebruikt", "Beschadigd"]

STAAT_DETAILS = {
    "Nieuwstaat": ["In verpakking", "Zonder verpakking", "Beschadigde verpakking"],
    "Zo goed als nieuw": ["Lichte gebruikerssporen", "Ongebruikt", "Goede staat"],
    "Gebruikt": ["Gebruikerssporen", "Kleine schades zoals krassen en vlekken", "Mogelijk ontbrekende onderdelen"],
    "Beschadigd": ["Ontbrekende onderdelen", "Zware gebruikerssporen", "Opvallende krassen, sporen en vlekken", "Werking onbekend"],
}

DEFAULT_CATEGORIEEN = [
    "Huishouden", "Elektra/Elektronica", "Meubels", "Verlichting", "Speelgoed",
    "Kleding/Textiel", "Boeken/Media", "Sieraden/Accessoires", "Sport & Vrije tijd",
    "Tuin & Buiten", "Gereedschap", "Servies/Keuken", "Kunst & Decoratie", "Overig",
]

DEFAULT_TIJDSPERIODES = [
    "Antiek (100+ jaar)", "Vintage (20-100 jaar)", "Retro", "Klassiek",
    "Modern/Hedendaags", "Onbekend",
]

MAX_TITEL_LENGTE = 60

TEXTVIEW_CSS = b"""
textview {
    border: 1px solid alpha(@borders, 0.6);
    border-radius: 4px;
}
textview text {
    background-color: #3a3a3a;
    color: #e8e8e8;
    padding: 5px;
}
"""


def apply_css():
    provider = Gtk.CssProvider()
    provider.load_from_data(TEXTVIEW_CSS)
    screen = Gdk.Screen.get_default()
    Gtk.StyleContext.add_provider_for_screen(
        screen, provider, Gtk.STYLE_PROVIDER_PRIORITY_APPLICATION
    )

ALGEMENE_VOORWAARDEN = """Algemene voorwaarden:
Wij zijn een kleine kringloopwinkel en proberen z.s.m. te reageren. Soms is het heel druk in de winkel en lukt dit niet dezelfde dag.

De richtprijs van een product baseren wij op bestaand aanbod en staat van het artikel met minimum/maximum schatting.

Graag alleen biedingen plaatsen via de bied optie van marktplaats. Advertenties laten we vaak minimaal 2 weken online staan voordat we akkoord gaan met het hoogste aannemelijke bod. Bedankt voor uw begrip 🙏🏻

Als wij akkoord gaan met uw bod, reserveren wij het product maximaal een week voor u. U kunt het product ophalen en afrekenen in onze winkel of kiezen voor verzending.

U bent altijd welkom in onze winkel, maar langskomen voor de Marktplaats advertenties zonder afspraak wordt niet op prijs gesteld. Dit gaat altijd via specifieke medewerkers.

Let op: Bij ophalen in de winkel vervalt het herroepingsrecht en kun je het product ter plekke testen.

Wij zijn een Stichting en onderdeel van Samen Circulair Hoeksche Waard. Alle inkomsten gaan naar de huur en naar het creëren van plekken voor dagbesteding voor mensen die speciale zorg onvangen o.a. via Pameijer, Cavent en Welzijn Hoeksche Waard. Wij maken geen winst!!!

Ons adres:
Kringloop HerNieuw
Willem Beukelszstraat 6B
3261 LV Oud-Beijerland

Openingstijden Kringloop:
Dinsdag t/m vrijdag 10.00 – 16.00 uur
Zaterdag 10:00-13:00 uur
Zondag en maandag Gesloten"""


def config_dir():
    return os.path.dirname(os.path.abspath(__file__))


def config_path():
    return os.path.join(config_dir(), "productmanager_config.json")


def users_path():
    return os.path.join(config_dir(), "productmanager_users.json")


def user_settings_path(gebruikersnaam):
    """Pad voor gebruikersspecifieke instellingen."""
    settings_dir = os.path.join(config_dir(), "user_settings")
    os.makedirs(settings_dir, exist_ok=True)
    return os.path.join(settings_dir, f"{gebruikersnaam}_settings.json")


# ============================================
# GEBRUIKERSBEHEER
# ============================================
def _hash_password(password, salt=None):
    if salt is None:
        salt = secrets.token_hex(16)
    digest = hashlib.sha256((salt + password).encode("utf-8")).hexdigest()
    return salt, digest


class UserManager:
    def __init__(self):
        self.path = users_path()
        self.data = self._load()

    def _load(self):
        if os.path.exists(self.path):
            try:
                with open(self.path, "r", encoding="utf-8") as f:
                    return json.load(f)
            except Exception:
                pass
        return {}

    def save(self):
        with open(self.path, "w", encoding="utf-8") as f:
            json.dump(self.data, f, indent=2, ensure_ascii=False)

    def heeft_gebruikers(self):
        return len(self.data) > 0

    def gebruikersnamen(self):
        return sorted(self.data.keys())

    def voeg_toe(self, gebruikersnaam, wachtwoord):
        salt, digest = _hash_password(wachtwoord)
        self.data[gebruikersnaam] = {"salt": salt, "hash": digest}
        self.save()

    def verwijder(self, gebruikersnaam):
        if gebruikersnaam in self.data:
            del self.data[gebruikersnaam]
            self.save()
        settings_path = user_settings_path(gebruikersnaam)
        if os.path.exists(settings_path):
            try:
                os.remove(settings_path)
            except Exception:
                pass

    def controleer(self, gebruikersnaam, wachtwoord):
        info = self.data.get(gebruikersnaam)
        if not info:
            return False
        _, digest = _hash_password(wachtwoord, info["salt"])
        return digest == info["hash"]


# ============================================
# GEBRUIKERSINSTELLINGEN
# ============================================
class UserSettings:
    """Slaat gebruikersspecifieke instellingen op zoals kolomvolgorde en zichtbare kolommen per filter."""

    def __init__(self, gebruikersnaam):
        self.gebruikersnaam = gebruikersnaam
        self.path = user_settings_path(gebruikersnaam)
        self.data = self._load()

    def _load(self):
        if os.path.exists(self.path):
            try:
                with open(self.path, "r", encoding="utf-8") as f:
                    return json.load(f)
            except Exception:
                pass
        return self._get_defaults()

    def _get_defaults(self):
        return {
            "zichtbare_kolommen": DEFAULT_ZICHTBARE_PER_FILTER.copy(),
            "kolom_volgorde": {},  # {filter: [kolomnamen]}
            "kolom_breedtes": {},  # {kolomnaam: breedte_in_pixels}
            "actieve_filter": FILTER_OFFLINE,
        }

    def save(self):
        with open(self.path, "w", encoding="utf-8") as f:
            json.dump(self.data, f, indent=2, ensure_ascii=False)

    def get(self, key, default=None):
        return self.data.get(key, default)

    def set(self, key, value):
        self.data[key] = value
        self.save()

    def get_zichtbare_kolommen(self, filter_type=FILTER_OFFLINE):
        return self.data.get("zichtbare_kolommen", {}).get(filter_type, DEFAULT_ZICHTBARE_PER_FILTER.get(filter_type, []))

    def set_zichtbare_kolommen(self, filter_type, kolommen):
        if "zichtbare_kolommen" not in self.data:
            self.data["zichtbare_kolommen"] = {}
        self.data["zichtbare_kolommen"][filter_type] = kolommen
        self.save()

    def get_kolom_volgorde(self, filter_type=FILTER_OFFLINE):
        return self.data.get("kolom_volgorde", {}).get(filter_type, [])

    def set_kolom_volgorde(self, filter_type, volgorde):
        if "kolom_volgorde" not in self.data:
            self.data["kolom_volgorde"] = {}
        self.data["kolom_volgorde"][filter_type] = volgorde
        self.save()

    def get_kolom_breedte(self, kolomnaam, default=100):
        return self.data.get("kolom_breedtes", {}).get(kolomnaam, default)

    def set_kolom_breedte(self, kolomnaam, breedte):
        if "kolom_breedtes" not in self.data:
            self.data["kolom_breedtes"] = {}
        self.data["kolom_breedtes"][kolomnaam] = breedte
        self.save()

    def get_actieve_filter(self):
        return self.data.get("actieve_filter", FILTER_OFFLINE)

    def set_actieve_filter(self, filter_type):
        self.data["actieve_filter"] = filter_type
        self.save()


# ============================================
# CONFIGURATIE
# ============================================
class ConfigManager:
    DEFAULTS = {
        "storage_backend": "xml",
        "xml_path": os.path.join(config_dir(), "producten.xml"),
        "sheets": {
            "sheet_url": "",
            "credentials_file": os.path.join(config_dir(), "credentials.json"),
        },
        "base_folder": os.path.expanduser("~/Documents/MarktplaatsProgramma/Producten"),
        "categorieen": DEFAULT_CATEGORIEEN,
        "tijdsperiodes": DEFAULT_TIJDSPERIODES,
        "opslaglocatie_codes": {"Locatie A": "A", "Locatie B": "B"},
        "sublocatie_codes": {"Sublocatie 1": "1", "Sublocatie 2": "2"},
        "rij_codes": {"Rij 1": "1", "Rij 2": "2"},
        "volgnummers": {},
        "algemene_voorwaarden_pad": "",  # Pad naar eigen algemene voorwaarden bestand
        "printer": {
            "enabled": False,
            "print_method": "cups",
            "printer_name": "",
            "brother_ql_model": "QL-500",
            "brother_ql_label": "62",
            "brother_ql_identifier": "",
            "auto_cut": True,
        },
        "barcode": {
            "module_width_mm": 0.4,
            "module_height_mm": 7.0,
            "quiet_zone_mm": 2.0,
        },
    }

    def __init__(self):
        self.path = config_path()
        self.data = self._load()

    def _load(self):
        if os.path.exists(self.path):
            try:
                with open(self.path, "r", encoding="utf-8") as f:
                    loaded = json.load(f)
                merged = json.loads(json.dumps(self.DEFAULTS))
                merged.update(loaded)
                return merged
            except Exception:
                pass
        return json.loads(json.dumps(self.DEFAULTS))

    def save(self):
        with open(self.path, "w", encoding="utf-8") as f:
            json.dump(self.data, f, indent=2, ensure_ascii=False)

    def get(self, key, default=None):
        return self.data.get(key, default)

    def set(self, key, value):
        self.data[key] = value
        self.save()


# ============================================
# OPSLAG-BACKENDS
# ============================================
class StorageBackend:
    def load_products(self):
        raise NotImplementedError

    def add_product(self, product):
        raise NotImplementedError

    def update_product(self, artikelnummer, product):
        raise NotImplementedError

    def delete_product(self, artikelnummer):
        raise NotImplementedError


class XmlBackend(StorageBackend):
    def __init__(self, xml_path):
        self.xml_path = xml_path

    def _ensure_file(self):
        if not os.path.exists(self.xml_path):
            import xml.etree.ElementTree as ET
            root = ET.Element("producten")
            tree = ET.ElementTree(root)
            os.makedirs(os.path.dirname(self.xml_path), exist_ok=True)
            tree.write(self.xml_path, encoding="utf-8", xml_declaration=True)

    def load_products(self):
        import xml.etree.ElementTree as ET
        self._ensure_file()
        tree = ET.parse(self.xml_path)
        root = tree.getroot()
        products = []
        for prod_el in root.findall("product"):
            product = {}
            for col in COLUMNS:
                el = prod_el.find(col)
                product[col] = el.text if el is not None and el.text else ""
            products.append(product)
        return products

    def _write_all(self, products):
        import xml.etree.ElementTree as ET
        root = ET.Element("producten")
        for product in products:
            prod_el = ET.SubElement(root, "product")
            for col in COLUMNS:
                el = ET.SubElement(prod_el, col)
                el.text = str(product.get(col, ""))
        tree = ET.ElementTree(root)
        os.makedirs(os.path.dirname(self.xml_path), exist_ok=True)
        tree.write(self.xml_path, encoding="utf-8", xml_declaration=True)

    def add_product(self, product):
        products = self.load_products()
        products.append(product)
        self._write_all(products)

    def update_product(self, artikelnummer, product):
        products = self.load_products()
        for i, p in enumerate(products):
            if p.get("artikelnummer") == artikelnummer:
                products[i] = product
                break
        self._write_all(products)

    def delete_product(self, artikelnummer):
        products = self.load_products()
        products = [p for p in products if p.get("artikelnummer") != artikelnummer]
        self._write_all(products)


class SheetsBackend(StorageBackend):
    def __init__(self, sheet_url, credentials_file):
        self.sheet_url = sheet_url
        self.credentials_file = credentials_file
        self._client = None
        self._worksheet = None

    def _connect(self):
        if self._worksheet is not None:
            return self._worksheet
        import gspread
        from google.oauth2.service_account import Credentials
        scopes = [
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive",
        ]
        creds = Credentials.from_service_account_file(self.credentials_file, scopes=scopes)
        self._client = gspread.authorize(creds)
        sheet = self._client.open_by_url(self.sheet_url)
        self._worksheet = sheet.get_worksheet(0)
        return self._worksheet

    def load_products(self):
        ws = self._connect()
        all_values = ws.get_all_values()
        products = []
        for row in all_values[1:]:
            if not row or not row[0]:
                continue
            product = {}
            for i, col in enumerate(COLUMNS):
                product[col] = row[i] if i < len(row) else ""
            products.append(product)
        return products

    def _product_to_row(self, product):
        return [str(product.get(col, "")) for col in COLUMNS]

    def add_product(self, product):
        ws = self._connect()
        ws.append_row(self._product_to_row(product))

    def _find_row_number(self, ws, artikelnummer):
        all_values = ws.get_all_values()
        for idx, row in enumerate(all_values, start=1):
            if row and row[0] == artikelnummer:
                return idx
        return None

    def update_product(self, artikelnummer, product):
        ws = self._connect()
        row_num = self._find_row_number(ws, artikelnummer)
        if row_num is None:
            self.add_product(product)
            return
        row = self._product_to_row(product)
        ws.update(f"A{row_num}:{chr(65 + len(COLUMNS) - 1)}{row_num}", [row])

    def delete_product(self, artikelnummer):
        ws = self._connect()
        row_num = self._find_row_number(ws, artikelnummer)
        if row_num is not None:
            ws.delete_rows(row_num)


def get_backend(config):
    backend = config.get("storage_backend", "xml")
    if backend == "sheets":
        sheets_cfg = config.get("sheets", {})
        return SheetsBackend(sheets_cfg.get("sheet_url", ""), sheets_cfg.get("credentials_file", ""))
    return XmlBackend(config.get("xml_path"))


# ============================================
# HELPERS
# ============================================
def generate_artikelnummer(config, storage, opslaglocatie_code, sublocatie_code, rij_code):
    prefix = f"{opslaglocatie_code}{sublocatie_code}{rij_code}"
    datum_deel = datetime.date.today().strftime("%d%m")
    teller_sleutel = f"{prefix}{datum_deel}"

    volgnummers = config.get("volgnummers", {})
    laatste = volgnummers.get(teller_sleutel, 0)

    try:
        bestaande = storage.load_products()
        for p in bestaande:
            nr = p.get("artikelnummer", "")
            if nr.startswith(teller_sleutel) and len(nr) == len(teller_sleutel) + 2:
                try:
                    val = int(nr[len(teller_sleutel):])
                    laatste = max(laatste, val)
                except ValueError:
                    pass
    except Exception:
        pass

    nieuw_nummer = laatste + 1
    if nieuw_nummer > 99:
        raise ValueError(
            f"Maximaal 99 artikelen per dag bereikt voor locatie '{prefix}' op {datum_deel}."
        )
    volgnummers[teller_sleutel] = nieuw_nummer
    config.set("volgnummers", volgnummers)

    return f"{teller_sleutel}{nieuw_nummer:02d}"


def build_txt_beschrijving(product, config=None):
    """Genereert omschrijving.txt in het mapje van het artikelnummer.
    Gebruikt de algemene voorwaarden uit het configuratiebestand indien aanwezig."""
    parts = []

    titel = product.get("titel", "")
    if titel:
        parts.append(titel)

    omschrijving = product.get("omschrijving", "")
    if omschrijving:
        parts.append(omschrijving)

    specs = []
    lengte = product.get("lengte", "")
    breedte = product.get("breedte", "")
    hoogte = product.get("hoogte", "")
    gewicht = product.get("gewicht", "")
    if lengte or breedte or hoogte or gewicht:
        dims = [f"{x}cm" for x in (lengte, breedte, hoogte) if x]
        line = f"Afmeting (LxBxH & G): {' x '.join(dims)}"
        if gewicht:
            line += f" & {gewicht} kg"
        specs.append(line)
    if product.get("conditie"):
        specs.append(f"Conditie/Staat: {product.get('conditie')}")
    if product.get("staat_details"):
        specs.append(f"Schades: {product.get('staat_details')}")
    waarde_min = product.get("waarde_min", "")
    waarde_max = product.get("waarde_max", "")
    if waarde_min:
        line = f"Waarde: {waarde_min}"
        if waarde_max:
            line += f" ~{waarde_max}"
        specs.append(line)
    if product.get("artikelnummer"):
        specs.append(f"Artikelnummer: {product.get('artikelnummer')}")
    if specs:
        parts.append("\n".join(specs))

    # Algemene voorwaarden - probeer eerst uit bestand, anders standaard
    voorwaarden = product.get("algemene_voorwaarden")
    if not voorwaarden and config:
        av_pad = config.get("algemene_voorwaarden_pad", "")
        if av_pad and os.path.exists(av_pad):
            try:
                with open(av_pad, "r", encoding="utf-8") as f:
                    voorwaarden = f.read()
            except Exception:
                pass

    if not voorwaarden:
        voorwaarden = ALGEMENE_VOORWAARDEN

    parts.append(voorwaarden)

    return "\n\n".join(parts)


def maak_product_map(config, product):
    """Maakt (of hergebruikt) de map voor dit artikelnummer."""
    base_folder = config.get("base_folder")
    artikelnummer = product["artikelnummer"]
    folder = os.path.join(base_folder, artikelnummer)
    bestond_al = os.path.exists(folder)
    os.makedirs(folder, exist_ok=True)

    txt_path = os.path.join(folder, "omschrijving.txt")
    with open(txt_path, "w", encoding="utf-8") as f:
        f.write(build_txt_beschrijving(product, config))

    return folder, bestond_al


def genereer_barcode(artikelnummer, output_folder, barcode_config=None, count_label=None):
    """Genereert een Code128-barcode-afbeelding met het artikelnummer.
    Optioneel kan een telling (bijv. "1/3") boven de barcode worden gezet."""
    try:
        import barcode
        from barcode.writer import ImageWriter
        from PIL import Image, ImageDraw, ImageFont
    except ImportError:
        raise RuntimeError(
            "python-barcode of PIL is niet geïnstalleerd. Installeer met:\n"
            "pip install python-barcode pillow --break-system-packages"
        )

    if not hasattr(Image, "ANTIALIAS"):
        Image.ANTIALIAS = Image.LANCZOS

    barcode_config = barcode_config or {}
    writer_options = {
        "module_width": barcode_config.get("module_width_mm", 0.4),
        "module_height": barcode_config.get("module_height_mm", 7.0),
        "quiet_zone": barcode_config.get("quiet_zone_mm", 2.0),
        "font_size": 8,
        "text_distance": 3.0,
        "dpi": 300,
    }

    code128 = barcode.get("code128", artikelnummer, writer=ImageWriter())
    output_path = os.path.join(output_folder, f"{artikelnummer}_barcode")
    saved_path = code128.save(output_path, options=writer_options)

    # Als er een count_label is, voeg deze toe boven de barcode
    if count_label:
        try:
            img = Image.open(saved_path)
            draw = ImageDraw.Draw(img)

            # Probeer een font te laden, gebruik standaard als niet beschikbaar
            try:
                font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 20)
            except:
                font = ImageFont.load_default()

            # Bereken tekst grootte
            bbox = draw.textbbox((0, 0), count_label, font=font)
            text_width = bbox[2] - bbox[0]
            text_height = bbox[3] - bbox[1]

            # Maak een nieuwe afbeelding met extra ruimte boven
            new_height = img.height + text_height + 20
            new_img = Image.new('RGB', (img.width, new_height), 'white')
            new_img.paste(img, (0, text_height + 10))

            # Teken de tekst gecentreerd bovenaan
            draw = ImageDraw.Draw(new_img)
            text_x = (img.width - text_width) // 2
            draw.text((text_x, 5), count_label, fill='black', font=font)

            new_img.save(saved_path)
        except Exception as e:
            print(f"Waarschuwing: Kon count_label niet toevoegen: {e}")

    return saved_path


def print_barcode(image_path, printer_config, auto_cut=True):
    """Print de barcode-afbeelding met optionele auto-cut."""
    methode = printer_config.get("print_method", "cups")

    if methode == "brother_ql":
        _print_barcode_brother_ql(image_path, printer_config, auto_cut)
    else:
        _print_barcode_cups(image_path, printer_config.get("printer_name", ""))


def _print_barcode_cups(image_path, printer_name):
    cmd = ["lp"]
    if printer_name:
        cmd += ["-d", printer_name]
    cmd += [image_path]
    result = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    if result.returncode != 0:
        raise RuntimeError(result.stderr.decode("utf-8", errors="ignore"))


def _print_barcode_brother_ql(image_path, printer_config, auto_cut=True):
    try:
        from brother_ql.conversion import convert
        from brother_ql.backends.helpers import send
        from brother_ql.raster import BrotherQLRaster
    except ImportError:
        raise RuntimeError(
            "brother_ql is niet geïnstalleerd. Installeer met:\n"
            "pip install brother_ql --break-system-packages"
        )

    model = printer_config.get("brother_ql_model", "QL-500")
    label_size = printer_config.get("brother_ql_label", "62")
    identifier = printer_config.get("brother_ql_identifier", "")
    if not identifier:
        raise RuntimeError(
            "Geen USB-identifier ingesteld voor brother_ql.\n"
            "Vind 'm met: python3 -m brother_ql discover"
        )

    from PIL import Image
    with Image.open(image_path) as img:
        qlr = BrotherQLRaster(model)
        qlr.exception_on_warning = True
        instructions = convert(
            qlr=qlr,
            images=[img],
            label=label_size,
            rotate="0",
            threshold=70.0,
            dither=False,
            compress=False,
            red=False,
            cut=auto_cut,  # Auto-cut inschakelen
        )
        send(instructions=instructions, printer_identifier=identifier, backend_identifier="pyusb", blocking=True)


def get_cups_printers():
    try:
        result = subprocess.run(["lpstat", "-p"], stdout=subprocess.PIPE, stderr=subprocess.PIPE, timeout=5)
        printers = []
        for line in result.stdout.decode("utf-8", errors="ignore").splitlines():
            if line.startswith("printer "):
                printers.append(line.split()[1])
        return printers
    except Exception:
        return []


# ============================================
# LOGIN & GEBRUIKERSBEHEER-DIALOGEN
# ============================================
class NieuweGebruikerDialog(Gtk.Dialog):
    def __init__(self, parent, titel="Nieuwe gebruiker"):
        super().__init__(title=titel, transient_for=parent, flags=0)
        self.set_default_size(320, 180)
        self.add_buttons(Gtk.STOCK_CANCEL, Gtk.ResponseType.CANCEL, Gtk.STOCK_OK, Gtk.ResponseType.OK)

        box = self.get_content_area()
        box.set_spacing(8)
        box.set_border_width(12)

        box.pack_start(Gtk.Label(label="Gebruikersnaam:", xalign=0), False, False, 0)
        self.naam_entry = Gtk.Entry()
        box.pack_start(self.naam_entry, False, False, 0)

        box.pack_start(Gtk.Label(label="Wachtwoord:", xalign=0), False, False, 0)
        self.wachtwoord_entry = Gtk.Entry()
        self.wachtwoord_entry.set_visibility(False)
        box.pack_start(self.wachtwoord_entry, False, False, 0)

        box.pack_start(Gtk.Label(label="Herhaal wachtwoord:", xalign=0), False, False, 0)
        self.wachtwoord2_entry = Gtk.Entry()
        self.wachtwoord2_entry.set_visibility(False)
        box.pack_start(self.wachtwoord2_entry, False, False, 0)

        self.foutlabel = Gtk.Label()
        self.foutlabel.set_markup("<span foreground='red'></span>")
        box.pack_start(self.foutlabel, False, False, 0)

        self.show_all()

    def get_invoer(self):
        return (
            self.naam_entry.get_text().strip(),
            self.wachtwoord_entry.get_text(),
            self.wachtwoord2_entry.get_text(),
        )

    def toon_fout(self, msg):
        self.foutlabel.set_markup(f"<span foreground='red'>{msg}</span>")


class LoginDialog(Gtk.Dialog):
    def __init__(self, user_manager):
        super().__init__(title="Inloggen - Marktplaats Product Manager")
        self.user_manager = user_manager
        self.set_default_size(320, 220)
        self.set_position(Gtk.WindowPosition.CENTER)

        box = self.get_content_area()
        box.set_spacing(10)
        box.set_border_width(20)

        titel = Gtk.Label()
        titel.set_markup("<b>Marktplaats Product Manager</b>")
        box.pack_start(titel, False, False, 0)

        box.pack_start(Gtk.Label(label="Gebruikersnaam:", xalign=0), False, False, 0)
        self.naam_combo = Gtk.ComboBoxText.new_with_entry()
        for naam in user_manager.gebruikersnamen():
            self.naam_combo.append_text(naam)
        box.pack_start(self.naam_combo, False, False, 0)

        box.pack_start(Gtk.Label(label="Wachtwoord:", xalign=0), False, False, 0)
        self.wachtwoord_entry = Gtk.Entry()
        self.wachtwoord_entry.set_visibility(False)
        self.wachtwoord_entry.connect("activate", lambda w: self.response(Gtk.ResponseType.OK))
        box.pack_start(self.wachtwoord_entry, False, False, 0)

        self.foutlabel = Gtk.Label()
        box.pack_start(self.foutlabel, False, False, 0)

        nieuw_btn = Gtk.Button(label="Nieuwe gebruiker aanmaken")
        nieuw_btn.connect("clicked", self._on_nieuwe_gebruiker)
        box.pack_start(nieuw_btn, False, False, 5)

        self.add_buttons(Gtk.STOCK_CANCEL, Gtk.ResponseType.CANCEL, "Inloggen", Gtk.ResponseType.OK)

        self.show_all()

    def get_naam(self):
        tekst = self.naam_combo.get_active_text()
        return tekst.strip() if tekst else ""

    def get_wachtwoord(self):
        return self.wachtwoord_entry.get_text()

    def toon_fout(self, msg):
        self.foutlabel.set_markup(f"<span foreground='red'>{msg}</span>")

    def _on_nieuwe_gebruiker(self, widget):
        dialog = NieuweGebruikerDialog(self)
        while True:
            response = dialog.run()
            if response != Gtk.ResponseType.OK:
                break
            naam, ww1, ww2 = dialog.get_invoer()
            if not naam:
                dialog.toon_fout("Vul een gebruikersnaam in.")
                continue
            if naam in self.user_manager.gebruikersnamen():
                dialog.toon_fout("Deze gebruikersnaam bestaat al.")
                continue
            if not ww1 or ww1 != ww2:
                dialog.toon_fout("Wachtwoorden komen niet overeen.")
                continue
            self.user_manager.voeg_toe(naam, ww1)
            self.naam_combo.append_text(naam)
            self.naam_combo.set_active(len(self.user_manager.gebruikersnamen()) - 1)
            break
        dialog.destroy()


class UsersDialog(Gtk.Dialog):
    def __init__(self, parent, user_manager, huidige_gebruiker):
        super().__init__(title="Gebruikers beheren", transient_for=parent, flags=0)
        self.user_manager = user_manager
        self.huidige_gebruiker = huidige_gebruiker
        self.set_default_size(320, 300)
        self.add_buttons(Gtk.STOCK_CLOSE, Gtk.ResponseType.CLOSE)

        box = self.get_content_area()
        box.set_spacing(8)
        box.set_border_width(12)

        self.store = Gtk.ListStore(str)
        self.tree_view = Gtk.TreeView(model=self.store)
        renderer = Gtk.CellRendererText()
        self.tree_view.append_column(Gtk.TreeViewColumn("Gebruikersnaam", renderer, text=0))
        scroll = Gtk.ScrolledWindow()
        scroll.set_size_request(-1, 180)
        scroll.add(self.tree_view)
        box.pack_start(scroll, True, True, 0)

        knoppen_row = Gtk.Box(spacing=5)
        toevoegen_btn = Gtk.Button(label="➕ Toevoegen")
        toevoegen_btn.connect("clicked", self._on_toevoegen)
        knoppen_row.pack_start(toevoegen_btn, False, False, 0)

        verwijder_btn = Gtk.Button(label="🗑️ Verwijderen")
        verwijder_btn.connect("clicked", self._on_verwijderen)
        knoppen_row.pack_start(verwijder_btn, False, False, 0)
        box.pack_start(knoppen_row, False, False, 0)

        self._herlaad()
        self.show_all()

    def _herlaad(self):
        self.store.clear()
        for naam in self.user_manager.gebruikersnamen():
            self.store.append([naam])

    def _on_toevoegen(self, widget):
        dialog = NieuweGebruikerDialog(self)
        while True:
            response = dialog.run()
            if response != Gtk.ResponseType.OK:
                break
            naam, ww1, ww2 = dialog.get_invoer()
            if not naam:
                dialog.toon_fout("Vul een gebruikersnaam in.")
                continue
            if naam in self.user_manager.gebruikersnamen():
                dialog.toon_fout("Deze gebruikersnaam bestaat al.")
                continue
            if not ww1 or ww1 != ww2:
                dialog.toon_fout("Wachtwoorden komen niet overeen.")
                continue
            self.user_manager.voeg_toe(naam, ww1)
            self._herlaad()
            break
        dialog.destroy()

    def _on_verwijderen(self, widget):
        selection = self.tree_view.get_selection()
        model, treeiter = selection.get_selected()
        if treeiter is None:
            return
        naam = model[treeiter][0]
        if naam == self.huidige_gebruiker:
            confirm = Gtk.MessageDialog(
                transient_for=self, flags=0, message_type=Gtk.MessageType.WARNING,
                buttons=Gtk.ButtonsType.OK, text="Je kan de gebruiker waarmee je nu bent ingelogd niet verwijderen."
            )
            confirm.run()
            confirm.destroy()
            return
        if len(self.user_manager.gebruikersnamen()) <= 1:
            confirm = Gtk.MessageDialog(
                transient_for=self, flags=0, message_type=Gtk.MessageType.WARNING,
                buttons=Gtk.ButtonsType.OK, text="Er moet minstens één gebruiker overblijven."
            )
            confirm.run()
            confirm.destroy()
            return
        confirm = Gtk.MessageDialog(
            transient_for=self, flags=0, message_type=Gtk.MessageType.QUESTION,
            buttons=Gtk.ButtonsType.YES_NO, text=f"Gebruiker '{naam}' verwijderen?"
        )
        response = confirm.run()
        confirm.destroy()
        if response == Gtk.ResponseType.YES:
            self.user_manager.verwijder(naam)
            self._herlaad()


# ============================================
# INSTELLINGEN-DIALOOG
# ============================================
class SettingsDialog(Gtk.Dialog):
    def __init__(self, parent, config):
        super().__init__(title="Instellingen", transient_for=parent, flags=0)
        self.config = config
        self.set_default_size(560, 800)
        self.add_buttons(Gtk.STOCK_CANCEL, Gtk.ResponseType.CANCEL, Gtk.STOCK_SAVE, Gtk.ResponseType.OK)

        box = self.get_content_area()
        box.set_spacing(10)
        box.set_border_width(15)

        scrolled = Gtk.ScrolledWindow()
        scrolled.set_policy(Gtk.PolicyType.NEVER, Gtk.PolicyType.AUTOMATIC)
        box.pack_start(scrolled, True, True, 0)

        inner = Gtk.Box(orientation=Gtk.Orientation.VERTICAL, spacing=12)
        inner.set_border_width(5)
        scrolled.add(inner)

        # --- Opslagmethode ---
        inner.pack_start(self._section_label("Opslagmethode (kies er één)"), False, False, 0)

        self.backend_combo = Gtk.ComboBoxText()
        self.backend_combo.append("xml", "Lokaal XML-bestand")
        self.backend_combo.append("sheets", "Google Sheets")
        self.backend_combo.set_active_id(config.get("storage_backend", "xml"))
        inner.pack_start(self.backend_combo, False, False, 0)

        warning = Gtk.Label()
        warning.set_markup("<small><i>⚠️ Zorg dat deze keuze overeenkomt met wat auto_marktplaats.py gebruikt - beide apps moeten dezelfde opslag lezen.</i></small>")
        warning.set_xalign(0)
        warning.set_line_wrap(True)
        inner.pack_start(warning, False, False, 0)

        inner.pack_start(Gtk.Label(label="XML-bestandspad:"), False, False, 0)
        self.xml_path_entry = Gtk.Entry()
        self.xml_path_entry.set_text(config.get("xml_path", ""))
        inner.pack_start(self.xml_path_entry, False, False, 0)

        inner.pack_start(Gtk.Label(label="Google Sheets URL:"), False, False, 0)
        self.sheet_url_entry = Gtk.Entry()
        self.sheet_url_entry.set_text(config.get("sheets", {}).get("sheet_url", ""))
        inner.pack_start(self.sheet_url_entry, False, False, 0)

        inner.pack_start(Gtk.Label(label="credentials.json pad:"), False, False, 0)
        self.creds_entry = Gtk.Entry()
        self.creds_entry.set_text(config.get("sheets", {}).get("credentials_file", ""))
        inner.pack_start(self.creds_entry, False, False, 0)

        inner.pack_start(Gtk.Separator(), False, False, 5)

        # --- Basis-map ---
        inner.pack_start(self._section_label("Basismap voor productmappen"), False, False, 0)
        base_warn = Gtk.Label()
        base_warn.set_markup("<small><i>⚠️ Zorg dat dit exact dezelfde basismap is als in marktplaats_manager.py, anders maakt die een nieuwe/andere map aan voor hetzelfde artikelnummer.</i></small>")
        base_warn.set_xalign(0)
        base_warn.set_line_wrap(True)
        inner.pack_start(base_warn, False, False, 0)

        self.base_folder_entry = Gtk.Entry()
        self.base_folder_entry.set_text(config.get("base_folder", ""))
        inner.pack_start(self.base_folder_entry, False, False, 0)

        inner.pack_start(Gtk.Separator(), False, False, 5)

        # --- Algemene voorwaarden ---
        inner.pack_start(self._section_label("Algemene voorwaarden"), False, False, 0)
        av_hint = Gtk.Label()
        av_hint.set_markup("<small><i>Selecteer een .txt bestand met uw eigen algemene voorwaarden. De opmaak uit het bestand wordt gebruikt. Laat leeg voor de standaard voorwaarden.</i></small>")
        av_hint.set_xalign(0)
        av_hint.set_line_wrap(True)
        inner.pack_start(av_hint, False, False, 0)

        av_row = Gtk.Box(spacing=5)
        self.av_entry = Gtk.Entry()
        av_pad = config.get("algemene_voorwaarden_pad", "")
        self.av_entry.set_text(av_pad)
        self.av_entry.set_placeholder_text("Pad naar algemene_voorwaarden.txt")
        self.av_entry.connect("changed", self._on_av_entry_changed)
        av_row.pack_start(self.av_entry, True, True, 0)

        av_browse_btn = Gtk.Button(label="📂 Bladeren")
        av_browse_btn.connect("clicked", self._browse_av_bestand)
        av_row.pack_start(av_browse_btn, False, False, 0)

        av_view_btn = Gtk.Button(label="👁️ Bekijk")
        av_view_btn.connect("clicked", self._view_av_bestand)
        av_row.pack_start(av_view_btn, False, False, 0)

        inner.pack_start(av_row, False, False, 0)

        # Preview
        self.av_preview = Gtk.Label()
        self.av_preview.set_xalign(0)
        self.av_preview.set_line_wrap(True)
        self.av_preview.set_selectable(True)
        self._update_av_preview()
        inner.pack_start(self.av_preview, False, False, 0)

        inner.pack_start(Gtk.Separator(), False, False, 5)

        # --- Categorieën ---
        inner.pack_start(self._section_label("Categorieën (één per regel)"), False, False, 0)
        self.categorieen_view = Gtk.TextView()
        self.categorieen_view.get_buffer().set_text("\n".join(config.get("categorieen", DEFAULT_CATEGORIEEN)))
        cat_scroll = Gtk.ScrolledWindow()
        cat_scroll.set_size_request(-1, 80)
        cat_scroll.add(self.categorieen_view)
        inner.pack_start(cat_scroll, False, False, 0)

        # --- Tijdsperiodes ---
        inner.pack_start(self._section_label("Tijdsperiodes (één per regel)"), False, False, 0)
        self.tijdsperiodes_view = Gtk.TextView()
        self.tijdsperiodes_view.get_buffer().set_text("\n".join(config.get("tijdsperiodes", DEFAULT_TIJDSPERIODES)))
        tp_scroll = Gtk.ScrolledWindow()
        tp_scroll.set_size_request(-1, 80)
        tp_scroll.add(self.tijdsperiodes_view)
        inner.pack_start(tp_scroll, False, False, 0)

        inner.pack_start(Gtk.Separator(), False, False, 5)

        # --- Opslaglocatie-codes ---
        inner.pack_start(self._section_label("Opslaglocatie-codes (naam=code)"), False, False, 0)
        self.opslaglocatie_view = Gtk.TextView()
        self.opslaglocatie_view.get_buffer().set_text(
            "\n".join(f"{k}={v}" for k, v in config.get("opslaglocatie_codes", {}).items())
        )
        loc_scroll = Gtk.ScrolledWindow()
        loc_scroll.set_size_request(-1, 60)
        loc_scroll.add(self.opslaglocatie_view)
        inner.pack_start(loc_scroll, False, False, 0)

        inner.pack_start(Gtk.Label(label="Sublocatie-codes (naam=code):"), False, False, 0)
        self.sublocatie_view = Gtk.TextView()
        self.sublocatie_view.get_buffer().set_text(
            "\n".join(f"{k}={v}" for k, v in config.get("sublocatie_codes", {}).items())
        )
        sub_scroll = Gtk.ScrolledWindow()
        sub_scroll.set_size_request(-1, 60)
        sub_scroll.add(self.sublocatie_view)
        inner.pack_start(sub_scroll, False, False, 0)

        inner.pack_start(Gtk.Label(label="Rij-codes (naam=code):"), False, False, 0)
        self.rij_view = Gtk.TextView()
        self.rij_view.get_buffer().set_text(
            "\n".join(f"{k}={v}" for k, v in config.get("rij_codes", {}).items())
        )
        rij_scroll = Gtk.ScrolledWindow()
        rij_scroll.set_size_request(-1, 60)
        rij_scroll.add(self.rij_view)
        inner.pack_start(rij_scroll, False, False, 0)

        inner.pack_start(Gtk.Separator(), False, False, 5)

        # --- Labelprinter ---
        inner.pack_start(self._section_label("Labelprinter (barcode)"), False, False, 0)
        printer_hint = Gtk.Label()
        printer_hint.set_markup("<small><i>Indien uitgeschakeld wordt alleen een barcode-afbeelding opgeslagen in de productmap.</i></small>")
        printer_hint.set_xalign(0)
        printer_hint.set_line_wrap(True)
        inner.pack_start(printer_hint, False, False, 0)

        self.printer_enabled_check = Gtk.CheckButton(label="Direct printen inschakelen")
        self.printer_enabled_check.set_active(config.get("printer", {}).get("enabled", False))
        inner.pack_start(self.printer_enabled_check, False, False, 0)

        self.auto_cut_check = Gtk.CheckButton(label="Auto-cut inschakelen (snijden na printen)")
        self.auto_cut_check.set_active(config.get("printer", {}).get("auto_cut", True))
        inner.pack_start(self.auto_cut_check, False, False, 0)

        printer_cfg = config.get("printer", {})

        method_row = Gtk.Box(spacing=5)
        method_row.pack_start(Gtk.Label(label="Methode:"), False, False, 0)
        self.print_method_combo = Gtk.ComboBoxText()
        self.print_method_combo.append("cups", "CUPS (lp) - algemeen")
        self.print_method_combo.append("brother_ql", "Brother QL direct via USB (brother_ql)")
        self.print_method_combo.set_active_id(printer_cfg.get("print_method", "cups"))
        self.print_method_combo.connect("changed", self._on_print_method_changed)
        method_row.pack_start(self.print_method_combo, False, False, 0)
        inner.pack_start(method_row, False, False, 0)

        self.cups_box = Gtk.Box(orientation=Gtk.Orientation.VERTICAL, spacing=3)
        inner.pack_start(self.cups_box, False, False, 0)

        printer_row = Gtk.Box(spacing=5)
        printer_row.pack_start(Gtk.Label(label="Printer:"), False, False, 0)
        self.printer_combo = Gtk.ComboBoxText()
        for p in get_cups_printers():
            self.printer_combo.append_text(p)
        huidige_printer = printer_cfg.get("printer_name", "")
        if huidige_printer:
            self.printer_combo.prepend_text(huidige_printer)
        self.printer_combo.set_active(0)
        printer_row.pack_start(self.printer_combo, True, True, 0)

        refresh_btn = Gtk.Button(label="🔄")
        refresh_btn.connect("clicked", self._refresh_printers)
        printer_row.pack_start(refresh_btn, False, False, 0)

        self.cups_box.pack_start(printer_row, False, False, 0)

        self.brother_ql_box = Gtk.Box(orientation=Gtk.Orientation.VERTICAL, spacing=3)
        inner.pack_start(self.brother_ql_box, False, False, 0)

        bq_grid = Gtk.Grid()
        bq_grid.set_column_spacing(10)
        bq_grid.set_row_spacing(5)
        self.brother_ql_box.pack_start(bq_grid, False, False, 0)

        bq_grid.attach(Gtk.Label(label="Model:", xalign=0), 0, 0, 1, 1)
        self.bq_model_entry = Gtk.Entry()
        self.bq_model_entry.set_text(printer_cfg.get("brother_ql_model", "QL-500"))
        bq_grid.attach(self.bq_model_entry, 1, 0, 1, 1)

        bq_grid.attach(Gtk.Label(label="Labelformaat:", xalign=0), 0, 1, 1, 1)
        self.bq_label_entry = Gtk.Entry()
        self.bq_label_entry.set_text(printer_cfg.get("brother_ql_label", "62"))
        bq_grid.attach(self.bq_label_entry, 1, 1, 1, 1)

        bq_grid.attach(Gtk.Label(label="USB-identifier:", xalign=0), 0, 2, 1, 1)
        self.bq_identifier_entry = Gtk.Entry()
        self.bq_identifier_entry.set_text(printer_cfg.get("brother_ql_identifier", ""))
        self.bq_identifier_entry.set_placeholder_text("bijv. usb://0x04f9:0x2015/000M6Z401370")
        bq_grid.attach(self.bq_identifier_entry, 1, 2, 1, 1)

        bq_find_btn = Gtk.Button(label="🔍 Zoek printer")
        bq_find_btn.connect("clicked", self._discover_brother_ql)
        self.brother_ql_box.pack_start(bq_find_btn, False, False, 0)

        self._on_print_method_changed(self.print_method_combo)

        # --- Barcode-afmetingen ---
        size_hint = Gtk.Label()
        size_hint.set_markup("<small><i>Afmetingen van de gegenereerde barcode-afbeelding</i></small>")
        size_hint.set_xalign(0)
        size_hint.set_line_wrap(True)
        size_hint.set_margin_top(8)
        inner.pack_start(size_hint, False, False, 0)

        barcode_cfg = config.get("barcode", {})

        size_grid = Gtk.Grid()
        size_grid.set_column_spacing(10)
        size_grid.set_row_spacing(5)
        inner.pack_start(size_grid, False, False, 0)

        size_grid.attach(Gtk.Label(label="Streepdikte (mm):", xalign=0), 0, 0, 1, 1)
        self.module_width_entry = Gtk.Entry()
        self.module_width_entry.set_text(str(barcode_cfg.get("module_width_mm", 0.4)))
        size_grid.attach(self.module_width_entry, 1, 0, 1, 1)

        size_grid.attach(Gtk.Label(label="Streephoogte (mm):", xalign=0), 0, 1, 1, 1)
        self.module_height_entry = Gtk.Entry()
        self.module_height_entry.set_text(str(barcode_cfg.get("module_height_mm", 7.0)))
        size_grid.attach(self.module_height_entry, 1, 1, 1, 1)

        size_grid.attach(Gtk.Label(label="Marge/quiet zone (mm):", xalign=0), 0, 2, 1, 1)
        self.quiet_zone_entry = Gtk.Entry()
        self.quiet_zone_entry.set_text(str(barcode_cfg.get("quiet_zone_mm", 2.0)))
        size_grid.attach(self.quiet_zone_entry, 1, 2, 1, 1)

        self.show_all()

    def _on_av_entry_changed(self, widget):
        self._update_av_preview()

    def _browse_av_bestand(self, widget):
        dialog = Gtk.FileChooserDialog(
            title="Kies algemene voorwaarden bestand",
            transient_for=self,
            action=Gtk.FileChooserAction.OPEN
        )
        dialog.add_buttons(Gtk.STOCK_CANCEL, Gtk.ResponseType.CANCEL, Gtk.STOCK_OPEN, Gtk.ResponseType.OK)

        filter_txt = Gtk.FileFilter()
        filter_txt.set_name("Tekstbestanden (*.txt)")
        filter_txt.add_pattern("*.txt")
        dialog.add_filter(filter_txt)

        filter_all = Gtk.FileFilter()
        filter_all.set_name("Alle bestanden")
        filter_all.add_pattern("*")
        dialog.add_filter(filter_all)

        response = dialog.run()
        if response == Gtk.ResponseType.OK:
            bestandspad = dialog.get_filename()
            self.av_entry.set_text(bestandspad)
            self._update_av_preview()
        dialog.destroy()

    def _view_av_bestand(self, widget):
        av_pad = self.av_entry.get_text().strip()
        if not av_pad or not os.path.exists(av_pad):
            self._toon_info_dialoog("Geen geldig bestand geselecteerd.\nGebruik de knop 'Bladeren' om een .txt bestand te kiezen.")
            return

        try:
            with open(av_pad, "r", encoding="utf-8") as f:
                inhoud = f.read()

            dialog = Gtk.Dialog(
                title="Algemene voorwaarden - voorbeeld",
                transient_for=self,
                flags=0
            )
            dialog.set_default_size(500, 400)
            dialog.add_buttons(Gtk.STOCK_CLOSE, Gtk.ResponseType.CLOSE)

            box = dialog.get_content_area()
            box.set_spacing(8)
            box.set_border_width(10)

            label = Gtk.Label(label=f"Bestand: {os.path.basename(av_pad)}")
            label.set_xalign(0)
            box.pack_start(label, False, False, 0)

            scrolled = Gtk.ScrolledWindow()
            scrolled.set_policy(Gtk.PolicyType.AUTOMATIC, Gtk.PolicyType.AUTOMATIC)
            scrolled.set_size_request(-1, 350)
            box.pack_start(scrolled, True, True, 0)

            textview = Gtk.TextView()
            textview.set_wrap_mode(Gtk.WrapMode.WORD)
            textview.get_buffer().set_text(inhoud)
            textview.set_editable(False)
            scrolled.add(textview)

            dialog.show_all()
            dialog.run()
            dialog.destroy()

        except Exception as e:
            self._toon_info_dialoog(f"Fout bij lezen bestand: {e}")

    def _update_av_preview(self):
        av_pad = self.av_entry.get_text().strip()
        if av_pad and os.path.exists(av_pad):
            try:
                with open(av_pad, "r", encoding="utf-8") as f:
                    inhoud = f.read()
                regels = inhoud.splitlines()[:5]
                preview = "\n".join(regels)
                if len(inhoud.splitlines()) > 5:
                    preview += "\n..."
                self.av_preview.set_markup(f"<small><i>📄 Voorbeeld uit bestand:</i></small>\n<tt>{GLib.markup_escape_text(preview)}</tt>")
            except Exception as e:
                self.av_preview.set_markup(f"<span foreground='red'>⚠️ Kan bestand niet lezen: {e}</span>")
        else:
            regels = ALGEMENE_VOORWAARDEN.splitlines()[:5]
            preview = "\n".join(regels) + "\n..."
            self.av_preview.set_markup(f"<small><i>📄 Standaard voorwaarden (ingebouwd):</i></small>\n<tt>{GLib.markup_escape_text(preview)}</tt>")

    def _toon_info_dialoog(self, msg):
        dialog = Gtk.MessageDialog(
            transient_for=self, flags=0,
            message_type=Gtk.MessageType.INFO,
            buttons=Gtk.ButtonsType.OK,
            text=msg
        )
        dialog.run()
        dialog.destroy()

    # De rest van de methods blijven hetzelfde...
    def _on_print_method_changed(self, widget):
        is_brother_ql = self.print_method_combo.get_active_id() == "brother_ql"
        self.cups_box.set_visible(not is_brother_ql)
        self.brother_ql_box.set_visible(is_brother_ql)

    def _discover_brother_ql(self, widget):
        try:
            result = subprocess.run(
                ["python3", "-m", "brother_ql", "discover"],
                stdout=subprocess.PIPE, stderr=subprocess.PIPE, timeout=10
            )
            output = result.stdout.decode("utf-8", errors="ignore").strip()
            if output:
                for line in output.splitlines():
                    if "usb://" in line:
                        self.bq_identifier_entry.set_text(line.strip())
                        break
                self._toon_info_dialoog(f"Gevonden:\n{output}")
            else:
                self._toon_info_dialoog("Geen brother_ql-printer gevonden.")
        except FileNotFoundError:
            self._toon_info_dialoog("brother_ql is niet geïnstalleerd.")
        except Exception as e:
            self._toon_info_dialoog(f"Fout: {e}")

    def _refresh_printers(self, widget):
        self.printer_combo.remove_all()
        for p in get_cups_printers():
            self.printer_combo.append_text(p)
        self.printer_combo.set_active(0)

    def _section_label(self, text):
        label = Gtk.Label()
        label.set_markup(f"<b>{text}</b>")
        label.set_xalign(0)
        return label

    def _parse_key_value_lines(self, textview):
        buf = textview.get_buffer()
        text = buf.get_text(buf.get_start_iter(), buf.get_end_iter(), True)
        result = {}
        for line in text.splitlines():
            line = line.strip()
            if "=" in line:
                k, v = line.split("=", 1)
                result[k.strip()] = v.strip()
        return result

    def _parse_lines(self, textview):
        buf = textview.get_buffer()
        text = buf.get_text(buf.get_start_iter(), buf.get_end_iter(), True)
        return [l.strip() for l in text.splitlines() if l.strip()]

    def apply_to_config(self):
        self.config.set("storage_backend", self.backend_combo.get_active_id())
        self.config.set("xml_path", self.xml_path_entry.get_text().strip())
        self.config.set("sheets", {
            "sheet_url": self.sheet_url_entry.get_text().strip(),
            "credentials_file": self.creds_entry.get_text().strip(),
        })
        self.config.set("base_folder", self.base_folder_entry.get_text().strip())
        self.config.set("categorieen", self._parse_lines(self.categorieen_view))
        self.config.set("tijdsperiodes", self._parse_lines(self.tijdsperiodes_view))
        self.config.set("opslaglocatie_codes", self._parse_key_value_lines(self.opslaglocatie_view))
        self.config.set("sublocatie_codes", self._parse_key_value_lines(self.sublocatie_view))
        self.config.set("rij_codes", self._parse_key_value_lines(self.rij_view))

        # Algemene voorwaarden pad opslaan
        self.config.set("algemene_voorwaarden_pad", self.av_entry.get_text().strip())

        printer_name = self.printer_combo.get_active_text() or ""
        self.config.set("printer", {
            "enabled": self.printer_enabled_check.get_active(),
            "print_method": self.print_method_combo.get_active_id() or "cups",
            "printer_name": printer_name,
            "brother_ql_model": self.bq_model_entry.get_text().strip(),
            "brother_ql_label": self.bq_label_entry.get_text().strip(),
            "brother_ql_identifier": self.bq_identifier_entry.get_text().strip(),
            "auto_cut": self.auto_cut_check.get_active(),
        })

        def _safe_float(entry, default):
            try:
                return float(entry.get_text().strip().replace(",", "."))
            except ValueError:
                return default

        self.config.set("barcode", {
            "module_width_mm": _safe_float(self.module_width_entry, 0.4),
            "module_height_mm": _safe_float(self.module_height_entry, 7.0),
            "quiet_zone_mm": _safe_float(self.quiet_zone_entry, 2.0),
        })


# ============================================
# TAB 1: REGISTREREN
# ============================================
class RegistreerTab(Gtk.Box):
    def __init__(self, main_window):
        super().__init__(orientation=Gtk.Orientation.VERTICAL, spacing=10)
        self.main_window = main_window
        self.config = main_window.config
        self.set_border_width(15)

        self.condition_checks = {}
        self.condition_box = None

        scrolled = Gtk.ScrolledWindow()
        scrolled.set_policy(Gtk.PolicyType.NEVER, Gtk.PolicyType.AUTOMATIC)
        self.pack_start(scrolled, True, True, 0)

        form = Gtk.Box(orientation=Gtk.Orientation.VERTICAL, spacing=10)
        form.set_border_width(5)
        scrolled.add(form)

        # Titel
        form.pack_start(self._label("Titel (max 60 tekens)"), False, False, 0)
        title_row = Gtk.Box(spacing=5)
        self.titel_entry = Gtk.Entry()
        self.titel_entry.set_max_length(MAX_TITEL_LENGTE)
        self.titel_entry.connect("changed", self._on_titel_changed)
        title_row.pack_start(self.titel_entry, True, True, 0)
        self.titel_counter = Gtk.Label(label=f"0/{MAX_TITEL_LENGTE}")
        title_row.pack_start(self.titel_counter, False, False, 0)
        form.pack_start(title_row, False, False, 0)

        # Omschrijving
        form.pack_start(self._label("Omschrijving"), False, False, 0)
        self.omschrijving_view = Gtk.TextView()
        self.omschrijving_view.set_wrap_mode(Gtk.WrapMode.WORD)
        omschrijving_scroll = Gtk.ScrolledWindow()
        omschrijving_scroll.set_size_request(-1, 100)
        omschrijving_scroll.add(self.omschrijving_view)
        form.pack_start(omschrijving_scroll, False, False, 0)

        # Categorie & Tijdsperiode
        cat_row = Gtk.Box(spacing=10)
        cat_col = Gtk.Box(orientation=Gtk.Orientation.VERTICAL, spacing=3)
        cat_col.pack_start(self._label("Categorie"), False, False, 0)
        self.categorie_combo = Gtk.ComboBoxText.new_with_entry()
        for c in self.config.get("categorieen", DEFAULT_CATEGORIEEN):
            self.categorie_combo.append_text(c)
        cat_col.pack_start(self.categorie_combo, False, False, 0)
        cat_row.pack_start(cat_col, True, True, 0)

        tp_col = Gtk.Box(orientation=Gtk.Orientation.VERTICAL, spacing=3)
        tp_col.pack_start(self._label("Tijdsperiode"), False, False, 0)
        self.tijdsperiode_combo = Gtk.ComboBoxText.new_with_entry()
        for t in self.config.get("tijdsperiodes", DEFAULT_TIJDSPERIODES):
            self.tijdsperiode_combo.append_text(t)
        tp_col.pack_start(self.tijdsperiode_combo, False, False, 0)
        cat_row.pack_start(tp_col, True, True, 0)
        form.pack_start(cat_row, False, False, 0)

        # Conditie
        form.pack_start(self._label("Staat van het artikel"), False, False, 0)
        self.conditie_combo = Gtk.ComboBoxText()
        for c in CONDITIES:
            self.conditie_combo.append_text(c)
        self.conditie_combo.connect("changed", self._on_conditie_changed)
        form.pack_start(self.conditie_combo, False, False, 0)

        self.condition_details_container = Gtk.Box(orientation=Gtk.Orientation.VERTICAL, spacing=3)
        form.pack_start(self.condition_details_container, False, False, 0)

        self.staat_details_extra_entry = Gtk.Entry()
        self.staat_details_extra_entry.set_placeholder_text("Overige staat-details (optioneel, vrije tekst)")
        form.pack_start(self.staat_details_extra_entry, False, False, 0)

        # Afmetingen & gewicht
        form.pack_start(self._label("Afmetingen (cm) en gewicht (kg)"), False, False, 0)
        dim_row = Gtk.Box(spacing=5)
        self.lengte_entry = self._number_entry("Lengte")
        self.breedte_entry = self._number_entry("Breedte")
        self.hoogte_entry = self._number_entry("Hoogte")
        self.gewicht_entry = self._number_entry("Gewicht")
        for e in (self.lengte_entry, self.breedte_entry, self.hoogte_entry, self.gewicht_entry):
            dim_row.pack_start(e, True, True, 0)
        form.pack_start(dim_row, False, False, 0)

        # Waarde & vraagprijs
        form.pack_start(self._label("Geschatte waarde en vraagprijs (€)"), False, False, 0)
        waarde_row = Gtk.Box(spacing=5)
        self.waarde_min_entry = self._number_entry("Min. waarde")
        self.waarde_max_entry = self._number_entry("Max. waarde")
        self.vraagprijs_entry = self._number_entry("Vraagprijs")
        for e in (self.waarde_min_entry, self.waarde_max_entry, self.vraagprijs_entry):
            waarde_row.pack_start(e, True, True, 0)
        form.pack_start(waarde_row, False, False, 0)

        # Opslaglocatie / Sublocatie / Rij
        form.pack_start(self._label("Opslaglocatie (bepaalt de eerste 3 tekens van het artikelnummer)"), False, False, 0)
        loc_row = Gtk.Box(spacing=5)
        self.opslaglocatie_combo = Gtk.ComboBoxText()
        for naam in self.config.get("opslaglocatie_codes", {}).keys():
            self.opslaglocatie_combo.append_text(naam)
        loc_row.pack_start(self.opslaglocatie_combo, True, True, 0)

        self.sublocatie_combo = Gtk.ComboBoxText()
        for naam in self.config.get("sublocatie_codes", {}).keys():
            self.sublocatie_combo.append_text(naam)
        loc_row.pack_start(self.sublocatie_combo, True, True, 0)

        self.rij_combo = Gtk.ComboBoxText()
        for naam in self.config.get("rij_codes", {}).keys():
            self.rij_combo.append_text(naam)
        loc_row.pack_start(self.rij_combo, True, True, 0)
        form.pack_start(loc_row, False, False, 0)

        warning_label = Gtk.Label()
        warning_label.set_markup(
            "<small><i>⚠️ Zorg dat de basismap-instelling overeenkomt met marktplaats_manager, "
            "zodat die hetzelfde mapje voor dit artikelnummer hergebruikt (de map bestaat dan al).</i></small>"
        )
        warning_label.set_xalign(0)
        warning_label.set_line_wrap(True)
        form.pack_start(warning_label, False, False, 5)

        # Barcode opties
        barcode_box = Gtk.Box(orientation=Gtk.Orientation.VERTICAL, spacing=5)
        barcode_box.set_border_width(5)
        barcode_box.set_margin_top(5)

        self.barcode_check = Gtk.CheckButton(label="Barcode genereren bij opslaan")
        self.barcode_check.set_active(True)
        barcode_box.pack_start(self.barcode_check, False, False, 0)

        self.meerdere_barcodes_check = Gtk.CheckButton(label="Meerdere barcodes printen (bijv. voor sets)")
        self.meerdere_barcodes_check.connect("toggled", self._on_meerdere_barcodes_toggled)
        barcode_box.pack_start(self.meerdere_barcodes_check, False, False, 0)

        # Aantal barcodes invoer (verborgen standaard)
        self.aantal_barcodes_row = Gtk.Box(spacing=5)
        self.aantal_barcodes_row.set_visible(False)
        self.aantal_barcodes_row.pack_start(Gtk.Label(label="Aantal barcodes:"), False, False, 0)
        self.aantal_barcodes_entry = Gtk.Entry()
        self.aantal_barcodes_entry.set_text("3")
        self.aantal_barcodes_entry.set_width_chars(5)
        self.aantal_barcodes_row.pack_start(self.aantal_barcodes_entry, False, False, 0)
        barcode_box.pack_start(self.aantal_barcodes_row, False, False, 0)

        form.pack_start(barcode_box, False, False, 0)

        # Opslaan-knop
        self.opslaan_btn = Gtk.Button(label="💾 Product registreren")
        self.opslaan_btn.connect("clicked", self._on_opslaan)
        form.pack_start(self.opslaan_btn, False, False, 10)

        self.status_label = Gtk.Label()
        self.status_label.set_xalign(0)
        self.status_label.set_line_wrap(True)
        form.pack_start(self.status_label, False, False, 0)

        self._on_conditie_changed(self.conditie_combo)

    def _on_meerdere_barcodes_toggled(self, widget):
        self.aantal_barcodes_row.set_visible(widget.get_active())

    def _label(self, text):
        label = Gtk.Label()
        label.set_markup(f"<b>{text}</b>")
        label.set_xalign(0)
        return label

    def _number_entry(self, placeholder):
        entry = Gtk.Entry()
        entry.set_placeholder_text(placeholder)
        return entry

    def _on_titel_changed(self, widget):
        length = len(widget.get_text())
        self.titel_counter.set_text(f"{length}/{MAX_TITEL_LENGTE}")

    def _on_conditie_changed(self, widget):
        for child in self.condition_details_container.get_children():
            self.condition_details_container.remove(child)
        self.condition_checks = {}

        conditie = widget.get_active_text()
        opties = STAAT_DETAILS.get(conditie, [])
        for optie in opties:
            check = Gtk.CheckButton(label=optie)
            self.condition_checks[optie] = check
            self.condition_details_container.pack_start(check, False, False, 0)
        self.condition_details_container.show_all()

    def _verzamel_staat_details(self):
        geselecteerd = [naam for naam, check in self.condition_checks.items() if check.get_active()]
        extra = self.staat_details_extra_entry.get_text().strip()
        if extra:
            geselecteerd.append(extra)
        return ", ".join(geselecteerd)

    def _reset_form(self):
        self.titel_entry.set_text("")
        self.omschrijving_view.get_buffer().set_text("")
        self.categorie_combo.get_child().set_text("")
        self.tijdsperiode_combo.get_child().set_text("")
        self.conditie_combo.set_active(0)
        self.staat_details_extra_entry.set_text("")
        for e in (self.lengte_entry, self.breedte_entry, self.hoogte_entry, self.gewicht_entry,
                  self.waarde_min_entry, self.waarde_max_entry, self.vraagprijs_entry):
            e.set_text("")
        self.meerdere_barcodes_check.set_active(False)
        self.aantal_barcodes_entry.set_text("3")

    def _on_opslaan(self, widget):
        titel = self.titel_entry.get_text().strip()
        if not titel:
            self._toon_fout("Titel is verplicht.")
            return
        if len(titel) > MAX_TITEL_LENGTE:
            self._toon_fout(f"Titel mag maximaal {MAX_TITEL_LENGTE} tekens zijn.")
            return

        opslaglocatie_naam = self.opslaglocatie_combo.get_active_text()
        sublocatie_naam = self.sublocatie_combo.get_active_text()
        rij_naam = self.rij_combo.get_active_text()
        if not (opslaglocatie_naam and sublocatie_naam and rij_naam):
            self._toon_fout("Kies een opslaglocatie, sublocatie en rij.")
            return

        opslaglocatie_code = self.config.get("opslaglocatie_codes", {}).get(opslaglocatie_naam, "X")
        sublocatie_code = self.config.get("sublocatie_codes", {}).get(sublocatie_naam, "X")
        rij_code = self.config.get("rij_codes", {}).get(rij_naam, "X")

        buf = self.omschrijving_view.get_buffer()
        omschrijving = buf.get_text(buf.get_start_iter(), buf.get_end_iter(), True)

        storage = get_backend(self.config)

        try:
            artikelnummer = generate_artikelnummer(
                self.config, storage, opslaglocatie_code, sublocatie_code, rij_code
            )
        except Exception as e:
            self._toon_fout(f"Kon artikelnummer niet genereren: {e}")
            return

        product = {col: "" for col in COLUMNS}
        product.update({
            "artikelnummer": artikelnummer,
            "titel": titel,
            "categorie": self.categorie_combo.get_active_text() or "",
            "omschrijving": omschrijving,
            "lengte": self.lengte_entry.get_text().strip(),
            "breedte": self.breedte_entry.get_text().strip(),
            "hoogte": self.hoogte_entry.get_text().strip(),
            "gewicht": self.gewicht_entry.get_text().strip(),
            "conditie": self.conditie_combo.get_active_text() or "",
            "staat_details": self._verzamel_staat_details(),
            "waarde_min": self.waarde_min_entry.get_text().strip(),
            "waarde_max": self.waarde_max_entry.get_text().strip(),
            "vraagprijs": self.vraagprijs_entry.get_text().strip(),
            "aanmaakdatum": datetime.date.today().isoformat(),
            "aanmaaktijd": datetime.datetime.now().strftime("%H:%M:%S"),
            "tijdsperiode": self.tijdsperiode_combo.get_active_text() or "",
            "opslaglocatie": opslaglocatie_naam,
            "sublocatie": sublocatie_naam,
            "rij": rij_naam,
            "verkocht": "nee",
            "online": "nee",
            "algemene_voorwaarden": "",
            "verwerkt_door": self.main_window.gebruikersnaam,
            "toegewezen_aan": "",
            "url_1": "",
            "url_2": "",
            "url_3": "",
            "url_4": "",
            "url_5": "",
        })

        try:
            folder, bestond_al = maak_product_map(self.config, product)
            product["folder_locatie"] = folder
        except Exception as e:
            self._toon_fout(f"Kon productmap niet aanmaken: {e}")
            return

        barcode_msg = ""
        if self.barcode_check.get_active():
            try:
                # Bepaal of we meerdere barcodes moeten genereren
                meerdere = self.meerdere_barcodes_check.get_active()
                if meerdere:
                    try:
                        aantal = int(self.aantal_barcodes_entry.get_text().strip())
                    except ValueError:
                        aantal = 3
                    if aantal < 1:
                        aantal = 1
                    if aantal > 99:
                        aantal = 99
                else:
                    aantal = 1

                barcode_paths = []
                for i in range(1, aantal + 1):
                    count_label = f"{i}/{aantal}" if meerdere else None
                    path = genereer_barcode(
                        artikelnummer, folder, self.config.get("barcode", {}), count_label
                    )
                    barcode_paths.append(path)

                    # Print direct als ingeschakeld
                    printer_cfg = self.config.get("printer", {})
                    if printer_cfg.get("enabled"):
                        auto_cut = printer_cfg.get("auto_cut", True)
                        # Alleen auto-cut op de laatste label
                        if i < aantal:
                            auto_cut = False
                        print_barcode(path, printer_cfg, auto_cut)

                if len(barcode_paths) == 1:
                    barcode_msg = f"\n📊 Barcode opgeslagen: {barcode_paths[0]}"
                else:
                    barcode_msg = f"\n📊 {len(barcode_paths)} barcodes opgeslagen in: {folder}"

                if printer_cfg.get("enabled"):
                    methode = printer_cfg.get("print_method", "cups")
                    doel = printer_cfg.get("printer_name") if methode == "cups" else "Brother QL (USB)"
                    barcode_msg += f"\n🖨️ Verstuurd naar printer '{doel}'"
                    if printer_cfg.get("auto_cut", True):
                        barcode_msg += " (met auto-cut)"

            except Exception as e:
                barcode_msg = f"\n⚠️ Barcode-fout: {e}"

        try:
            storage.add_product(product)
        except Exception as e:
            self._toon_fout(f"Kon niet opslaan naar opslag-backend: {e}")
            return

        map_msg = "Bestaande map hergebruikt" if bestond_al else "Nieuwe map aangemaakt"
        self.status_label.set_markup(
            f"<span foreground='green'>✅ Product <b>{artikelnummer}</b> geregistreerd.\n"
            f"📁 {map_msg}: {folder}{barcode_msg}</span>"
        )
        self._reset_form()
        self.main_window.overzicht_tab.herlaad()
        self.main_window.notebook.set_current_page(1)

    def _toon_fout(self, msg):
        self.status_label.set_markup(f"<span foreground='red'>❌ {msg}</span>")

# ============================================
# TAB 2: OVERZICHT
# ============================================
class OverzichtTab(Gtk.Box):
    def __init__(self, main_window):
        super().__init__(orientation=Gtk.Orientation.VERTICAL, spacing=10)
        self.main_window = main_window
        self.config = main_window.config
        self.user_settings = main_window.user_settings
        self.set_border_width(15)

        # Huidige filter
        self.huidige_filter = self.user_settings.get_actieve_filter()
        self.zoekterm = ""

        # Filter-knoppen
        filter_row = Gtk.Box(spacing=5)
        self.filter_alles_btn = Gtk.RadioButton.new_with_label_from_widget(None, "Alles")
        self.filter_offline_btn = Gtk.RadioButton.new_with_label_from_widget(self.filter_alles_btn, "Offline")
        self.filter_online_btn = Gtk.RadioButton.new_with_label_from_widget(self.filter_alles_btn, "Online")
        self.filter_verkocht_btn = Gtk.RadioButton.new_with_label_from_widget(self.filter_alles_btn, "Verkocht")

        # Stel de actieve filter in
        if self.huidige_filter == FILTER_ALLES:
            self.filter_alles_btn.set_active(True)
        elif self.huidige_filter == FILTER_ONLINE:
            self.filter_online_btn.set_active(True)
        elif self.huidige_filter == FILTER_VERKOCHT:
            self.filter_verkocht_btn.set_active(True)
        else:
            self.filter_offline_btn.set_active(True)

        for btn, filter_type in [
            (self.filter_alles_btn, FILTER_ALLES),
            (self.filter_offline_btn, FILTER_OFFLINE),
            (self.filter_online_btn, FILTER_ONLINE),
            (self.filter_verkocht_btn, FILTER_VERKOCHT)
        ]:
            btn.connect("toggled", self._on_filter_changed, filter_type)
            filter_row.pack_start(btn, False, False, 0)

        # Zoekveld
        self.zoek_entry = Gtk.Entry()
        self.zoek_entry.set_placeholder_text("🔍 Zoek in artikelnummer of titel...")
        self.zoek_entry.connect("changed", self._on_zoek_changed)
        filter_row.pack_start(self.zoek_entry, True, True, 5)

        refresh_btn = Gtk.Button(label="🔄 Vernieuwen")
        refresh_btn.connect("clicked", lambda w: self.herlaad())
        filter_row.pack_start(refresh_btn, False, False, 0)

        self.pack_start(filter_row, False, False, 0)

        # Extra knoppen rij
        extra_row = Gtk.Box(spacing=5)

        # Kolomkiezer knop
        kolom_btn = Gtk.Button(label="📋 Kolommen kiezen")
        kolom_btn.connect("clicked", self._open_kolomkiezer)
        extra_row.pack_start(kolom_btn, False, False, 0)

        # Opslaan kolominstellingen knop
        save_col_btn = Gtk.Button(label="💾 Kolominstellingen opslaan")
        save_col_btn.connect("clicked", self._save_column_settings)
        extra_row.pack_start(save_col_btn, False, False, 0)

        # Exporteer knoppen
        export_csv_btn = Gtk.Button(label="📊 Exporteer CSV")
        export_csv_btn.connect("clicked", self._export_csv)
        extra_row.pack_start(export_csv_btn, False, False, 0)

        export_excel_btn = Gtk.Button(label="📊 Exporteer Excel")
        export_excel_btn.connect("clicked", self._export_excel)
        extra_row.pack_start(export_excel_btn, False, False, 0)

        # Dashboard knop
        dashboard_btn = Gtk.Button(label="📈 Dashboard")
        dashboard_btn.connect("clicked", self._open_dashboard)
        extra_row.pack_start(dashboard_btn, False, False, 0)

        self.pack_start(extra_row, False, False, 5)

        # Tabel - met meerdere selectie mogelijk
        self.store = Gtk.ListStore(*([str] * len(COLUMNS)))
        self.tree_view = Gtk.TreeView(model=self.store)
        self.tree_view.set_search_column(0)
        self.tree_view.get_selection().set_mode(Gtk.SelectionMode.MULTIPLE)
        self.tree_view.connect("row-activated", self._on_row_dubbelklik)
        self.tree_view.connect("columns-changed", self._on_columns_changed)

        # Maak kolommen op basis van gebruikersinstellingen
        self._update_tree_columns()

        scrolled = Gtk.ScrolledWindow()
        scrolled.set_policy(Gtk.PolicyType.AUTOMATIC, Gtk.PolicyType.AUTOMATIC)
        scrolled.add(self.tree_view)
        self.pack_start(scrolled, True, True, 0)

        # Actie-knoppen
        actie_row = Gtk.Box(spacing=5)
        self.bewerk_btn = Gtk.Button(label="✏️ Bewerken")
        self.bewerk_btn.connect("clicked", self._on_bewerken)
        actie_row.pack_start(self.bewerk_btn, False, False, 0)

        self.verwijder_btn = Gtk.Button(label="🗑️ Verwijderen")
        self.verwijder_btn.connect("clicked", self._on_verwijderen)
        actie_row.pack_start(self.verwijder_btn, False, False, 0)

        self.online_btn = Gtk.Button(label="🌐 Markeer als online")
        self.online_btn.connect("clicked", self._on_markeer_online)
        actie_row.pack_start(self.online_btn, False, False, 0)

        self.offline_btn = Gtk.Button(label="📴 Markeer als offline")
        self.offline_btn.connect("clicked", self._on_markeer_offline)
        actie_row.pack_start(self.offline_btn, False, False, 0)

        self.verkocht_btn = Gtk.Button(label="💰 Markeer als verkocht")
        self.verkocht_btn.connect("clicked", self._on_markeer_verkocht)
        actie_row.pack_start(self.verkocht_btn, False, False, 0)

        self.toewijs_btn = Gtk.Button(label="👤 Toewijzen aan medewerker")
        self.toewijs_btn.connect("clicked", self._on_toewijzen)
        actie_row.pack_start(self.toewijs_btn, False, False, 0)

        # Barcode printen knop
        self.print_barcode_btn = Gtk.Button(label="🏷️ Print barcode(s)")
        self.print_barcode_btn.connect("clicked", self._on_print_barcodes)
        actie_row.pack_start(self.print_barcode_btn, False, False, 0)

        self.pack_start(actie_row, False, False, 0)

        # Status / Omzet label
        self.omzet_label = Gtk.Label()
        self.omzet_label.set_xalign(0)
        self.omzet_label.set_line_wrap(True)
        self.pack_start(self.omzet_label, False, False, 0)

        self.herlaad()

    def _on_filter_changed(self, button, filter_type):
        if button.get_active():
            self.huidige_filter = filter_type
            self.user_settings.set_actieve_filter(filter_type)
            self._update_tree_columns()
            self.herlaad()

    def _on_zoek_changed(self, widget):
        self.zoekterm = widget.get_text().strip().lower()
        self.herlaad()

    def _update_tree_columns(self):
        # Verwijder alle bestaande kolommen
        for col in self.tree_view.get_columns():
            self.tree_view.remove_column(col)

        # Bepaal de kolomvolgorde voor deze filter
        volgorde = self.user_settings.get_kolom_volgorde(self.huidige_filter)
        zichtbaar = self.user_settings.get_zichtbare_kolommen(self.huidige_filter)

        # Als er geen opgeslagen volgorde is, gebruik de zichtbare kolommen in de standaard volgorde
        if not volgorde:
            volgorde = [k for k in COLUMNS if k in zichtbaar]
        else:
            volgorde = [k for k in volgorde if k in zichtbaar]
            for k in zichtbaar:
                if k not in volgorde:
                    volgorde.append(k)

        for kolom in volgorde:
            if kolom in COLUMNS:
                idx = COLUMNS.index(kolom)
                label = KOLOM_LABELS.get(kolom, kolom)
                renderer = Gtk.CellRendererText()
                column = Gtk.TreeViewColumn(label, renderer, text=idx)
                column.set_sort_column_id(idx)
                column.set_resizable(True)
                column.set_reorderable(True)
                breedte = self.user_settings.get_kolom_breedte(kolom, 100)
                column.set_fixed_width(breedte)
                column.set_min_width(50)
                self.tree_view.append_column(column)

    def _on_columns_changed(self, widget):
        pass

    def _save_column_settings(self, widget):
        """Sla de huidige kolomvolgorde en breedtes op voor de huidige filter."""
        columns = self.tree_view.get_columns()
        volgorde = []
        breedtes = {}

        for col in columns:
            title = col.get_title()
            for kolom, label in KOLOM_LABELS.items():
                if label == title:
                    volgorde.append(kolom)
                    breedtes[kolom] = col.get_width()
                    break

        self.user_settings.set_kolom_volgorde(self.huidige_filter, volgorde)

        for kolom, breedte in breedtes.items():
            self.user_settings.set_kolom_breedte(kolom, breedte)

        self.omzet_label.set_markup(
            f"<span foreground='green'>✅ Kolominstellingen opgeslagen voor filter '{FILTER_NAMEN[self.huidige_filter]}'</span>"
        )
        GLib.timeout_add(3000, self._herstel_omzet_label)

    def _open_kolomkiezer(self, widget):
        dialog = KolomKiezerDialog(self.main_window, self.user_settings, self.huidige_filter)
        response = dialog.run()
        if response == Gtk.ResponseType.OK:
            dialog.apply_to_settings()
            self._update_tree_columns()
            self.herlaad()
        dialog.destroy()

    def _huidige_producten(self):
        storage = get_backend(self.config)
        try:
            return storage.load_products()
        except Exception as e:
            self._toon_foutdialoog(f"Kon producten niet laden: {e}")
            return []

    def herlaad(self):
        self.store.clear()
        producten = self._huidige_producten()

        # Filter op basis van status
        if self.huidige_filter == FILTER_OFFLINE:
            producten = [p for p in producten
                         if p.get("verkocht", "nee").lower() != "ja"
                         and p.get("online", "nee").lower() != "ja"]
        elif self.huidige_filter == FILTER_ONLINE:
            producten = [p for p in producten
                         if p.get("verkocht", "nee").lower() != "ja"
                         and p.get("online", "nee").lower() == "ja"]
        elif self.huidige_filter == FILTER_VERKOCHT:
            producten = [p for p in producten if p.get("verkocht", "nee").lower() == "ja"]

        # Zoekfilter
        if self.zoekterm:
            producten = [p for p in producten
                         if self.zoekterm in p.get("artikelnummer", "").lower()
                         or self.zoekterm in p.get("titel", "").lower()]

        for p in producten:
            row = [p.get(col, "") for col in COLUMNS]
            self.store.append(row)

        # Toon aantal gevonden producten
        if self.zoekterm:
            self.omzet_label.set_markup(f"<small>🔍 {len(producten)} resultaten gevonden voor '{self.zoekterm}'</small>")
        elif self.huidige_filter == FILTER_VERKOCHT:
            totaal = 0.0
            for p in producten:
                try:
                    prijs = p.get("verkoopprijs", "").replace("€", "").replace(",", ".").strip()
                    totaal += float(prijs) if prijs else 0.0
                except ValueError:
                    pass
            self.omzet_label.set_markup(
                f"<b>Totale omzet (verkocht, {len(producten)} items): €{totaal:.2f}</b>\n"
                f"<small><i>Let op: er wordt geen inkoopprijs bijgehouden (kringloop-donaties), "
                f"dus omzet = winst in dit overzicht.</i></small>"
            )
        else:
            self.omzet_label.set_text("")

    def _geselecteerde_artikelnummers(self):
        selection = self.tree_view.get_selection()
        model, paths = selection.get_selected_rows()
        artikelnummers = []
        for path in paths:
            treeiter = model.get_iter(path)
            if treeiter:
                artikelnummers.append(model[treeiter][0])
        return artikelnummers

    def _on_row_dubbelklik(self, tree_view, path, column):
        artikelnummer = self.store[path][0]
        klembord = Gtk.Clipboard.get(Gdk.SELECTION_CLIPBOARD)
        klembord.set_text(artikelnummer, -1)
        klembord.store()
        self.omzet_label.set_markup(f"<small>📋 Artikelnummer <b>{artikelnummer}</b> gekopieerd naar klembord</small>")
        GLib.timeout_add(1500, self._herstel_omzet_label)

    def _herstel_omzet_label(self):
        self.herlaad()
        return False

    def _toon_foutdialoog(self, msg):
        dialog = Gtk.MessageDialog(
            transient_for=self.main_window, flags=0,
            message_type=Gtk.MessageType.ERROR, buttons=Gtk.ButtonsType.OK, text=msg
        )
        dialog.run()
        dialog.destroy()

    def _toon_info_dialoog(self, msg, title="Informatie"):
        dialog = Gtk.MessageDialog(
            transient_for=self.main_window, flags=0,
            message_type=Gtk.MessageType.INFO, buttons=Gtk.ButtonsType.OK, text=msg
        )
        dialog.set_title(title)
        dialog.run()
        dialog.destroy()

    def _get_first_selected(self):
        nummers = self._geselecteerde_artikelnummers()
        return nummers[0] if nummers else None

    def _on_bewerken(self, widget):
        artikelnummer = self._get_first_selected()
        if not artikelnummer:
            self._toon_foutdialoog("Selecteer eerst een product.")
            return

        storage = get_backend(self.config)
        producten = storage.load_products()
        product = next((p for p in producten if p.get("artikelnummer") == artikelnummer), None)
        if not product:
            self._toon_foutdialoog("Product niet gevonden.")
            return

        dialog = BewerkDialog(self.main_window, product, self.config)
        response = dialog.run()
        if response == Gtk.ResponseType.OK:
            bijgewerkt = dialog.get_product()
            try:
                storage.update_product(artikelnummer, bijgewerkt)
                self.herlaad()
            except Exception as e:
                self._toon_foutdialoog(f"Kon niet opslaan: {e}")
        dialog.destroy()

    def _on_verwijderen(self, widget):
        artikelnummers = self._geselecteerde_artikelnummers()
        if not artikelnummers:
            self._toon_foutdialoog("Selecteer eerst een of meerdere producten.")
            return

        if len(artikelnummers) == 1:
            msg = f"Product {artikelnummers[0]} verwijderen?"
        else:
            msg = f"{len(artikelnummers)} producten verwijderen?"

        confirm = Gtk.MessageDialog(
            transient_for=self.main_window, flags=0,
            message_type=Gtk.MessageType.QUESTION, buttons=Gtk.ButtonsType.YES_NO,
            text=msg
        )
        if len(artikelnummers) > 1:
            confirm.format_secondary_text(f"Weet je zeker dat je deze {len(artikelnummers)} producten wilt verwijderen?")
        response = confirm.run()
        confirm.destroy()

        if response == Gtk.ResponseType.YES:
            storage = get_backend(self.config)
            for artikelnummer in artikelnummers:
                try:
                    producten = storage.load_products()
                    product = next((p for p in producten if p.get("artikelnummer") == artikelnummer), None)
                    folder = product.get("folder_locatie", "").strip() if product else ""

                    storage.delete_product(artikelnummer)

                    if folder and os.path.isdir(folder):
                        try:
                            from send2trash import send2trash
                            send2trash(folder)
                        except ImportError:
                            pass
                        except Exception:
                            pass
                except Exception as e:
                    self._toon_foutdialoog(f"Fout bij verwijderen {artikelnummer}: {e}")
                    return

            self.herlaad()

    def _on_markeer_online(self, widget):
        artikelnummers = self._geselecteerde_artikelnummers()
        if not artikelnummers:
            self._toon_foutdialoog("Selecteer eerst een of meerdere producten.")
            return

        dialog = OnlineDialog(self.main_window, artikelnummers)
        response = dialog.run()
        urls = dialog.get_urls() if response == Gtk.ResponseType.OK else None
        dialog.destroy()

        if urls is None:
            return

        storage = get_backend(self.config)
        producten = storage.load_products()

        for artikelnummer in artikelnummers:
            product = next((p for p in producten if p.get("artikelnummer") == artikelnummer), None)
            if not product:
                continue

            product["online"] = "ja"
            for i in range(1, 6):
                if i <= len(urls):
                    product[f"url_{i}"] = urls[i-1]
                else:
                    product[f"url_{i}"] = ""

            try:
                storage.update_product(artikelnummer, product)
            except Exception as e:
                self._toon_foutdialoog(f"Fout bij updaten {artikelnummer}: {e}")
                return

        self.herlaad()
        self._toon_info_dialoog(f"{len(artikelnummers)} product(en) gemarkeerd als online.")

    def _on_markeer_offline(self, widget):
        artikelnummers = self._geselecteerde_artikelnummers()
        if not artikelnummers:
            self._toon_foutdialoog("Selecteer eerst een of meerdere producten.")
            return

        if len(artikelnummers) == 1:
            msg = f"Markeer {artikelnummers[0]} als offline?"
        else:
            msg = f"Markeer {len(artikelnummers)} producten als offline?"

        confirm = Gtk.MessageDialog(
            transient_for=self.main_window, flags=0,
            message_type=Gtk.MessageType.QUESTION, buttons=Gtk.ButtonsType.YES_NO,
            text=msg
        )
        if len(artikelnummers) > 1:
            confirm.format_secondary_text(f"Alle advertentie-URLs worden verwijderd.")
        response = confirm.run()
        confirm.destroy()

        if response == Gtk.ResponseType.YES:
            storage = get_backend(self.config)
            producten = storage.load_products()

            for artikelnummer in artikelnummers:
                product = next((p for p in producten if p.get("artikelnummer") == artikelnummer), None)
                if not product:
                    continue

                product["online"] = "nee"
                for i in range(1, 6):
                    product[f"url_{i}"] = ""

                try:
                    storage.update_product(artikelnummer, product)
                except Exception as e:
                    self._toon_foutdialoog(f"Fout bij updaten {artikelnummer}: {e}")
                    return

            self.herlaad()
            self._toon_info_dialoog(f"{len(artikelnummers)} product(en) gemarkeerd als offline.")

    def _on_markeer_verkocht(self, widget):
        artikelnummers = self._geselecteerde_artikelnummers()
        if not artikelnummers:
            self._toon_foutdialoog("Selecteer eerst een of meerdere producten.")
            return

        if len(artikelnummers) > 1:
            self._toon_foutdialoog("Markeer als verkocht kan alleen voor één product tegelijk.")
            return

        artikelnummer = artikelnummers[0]

        dialog = VerkochtDialog(self.main_window, artikelnummer)
        response = dialog.run()
        gegevens = dialog.get_gegevens() if response == Gtk.ResponseType.OK else None
        dialog.destroy()

        if gegevens is None:
            return

        if not gegevens["verkoopprijs"]:
            self._toon_foutdialoog("Vul een verkoopprijs in.")
            return

        storage = get_backend(self.config)
        producten = storage.load_products()
        product = next((p for p in producten if p.get("artikelnummer") == artikelnummer), None)
        if not product:
            self._toon_foutdialoog("Product niet gevonden.")
            return

        product["verkocht"] = "ja"
        product.update(gegevens)
        try:
            storage.update_product(artikelnummer, product)
            self.herlaad()
            self._toon_info_dialoog(f"Product {artikelnummer} gemarkeerd als verkocht.")
        except Exception as e:
            self._toon_foutdialoog(f"Kon niet opslaan: {e}")

    def _on_toewijzen(self, widget):
        artikelnummers = self._geselecteerde_artikelnummers()
        if not artikelnummers:
            self._toon_foutdialoog("Selecteer eerst een of meerdere producten.")
            return

        storage = get_backend(self.config)
        producten = storage.load_products()
        eerste_product = next((p for p in producten if p.get("artikelnummer") == artikelnummers[0]), None)
        if not eerste_product:
            self._toon_foutdialoog("Product niet gevonden.")
            return

        dialog = ToewijzenDialog(self.main_window, eerste_product, self.main_window.user_manager, self.main_window.gebruikersnaam)
        response = dialog.run()
        if response == Gtk.ResponseType.OK:
            toegewezen_aan = dialog.get_toegewezen_aan()
            for artikelnummer in artikelnummers:
                product = next((p for p in producten if p.get("artikelnummer") == artikelnummer), None)
                if product:
                    product["toegewezen_aan"] = toegewezen_aan
                    try:
                        storage.update_product(artikelnummer, product)
                    except Exception as e:
                        self._toon_foutdialoog(f"Fout bij updaten {artikelnummer}: {e}")
                        return
            self.herlaad()
            self._toon_info_dialoog(f"{len(artikelnummers)} product(en) toegewezen aan {toegewezen_aan or 'niemand'}.")
        dialog.destroy()

    def _on_print_barcodes(self, widget):
        """Print barcodes voor geselecteerde producten."""
        artikelnummers = self._geselecteerde_artikelnummers()
        if not artikelnummers:
            self._toon_foutdialoog("Selecteer eerst een of meerdere producten.")
            return

        # Vraag of ze meerdere barcodes per product willen
        dialog = Gtk.Dialog(
            title="Barcodes printen",
            transient_for=self.main_window,
            flags=0
        )
        dialog.add_buttons(Gtk.STOCK_CANCEL, Gtk.ResponseType.CANCEL, "Printen", Gtk.ResponseType.OK)

        box = dialog.get_content_area()
        box.set_spacing(8)
        box.set_border_width(10)

        box.pack_start(Gtk.Label(label=f"Print barcodes voor {len(artikelnummers)} product(en):", xalign=0), False, False, 0)

        # Optie voor meerdere barcodes per product
        multi_check = Gtk.CheckButton(label="Meerdere barcodes per product (voor sets)")
        multi_check.set_active(False)
        box.pack_start(multi_check, False, False, 0)

        # Aantal invoer (verborgen)
        count_row = Gtk.Box(spacing=5)
        count_row.set_visible(False)
        count_row.pack_start(Gtk.Label(label="Aantal per product:"), False, False, 0)
        count_entry = Gtk.Entry()
        count_entry.set_text("3")
        count_entry.set_width_chars(5)
        count_row.pack_start(count_entry, False, False, 0)
        box.pack_start(count_row, False, False, 0)

        multi_check.connect("toggled", lambda w: count_row.set_visible(w.get_active()))

        dialog.show_all()
        response = dialog.run()

        meerdere = multi_check.get_active()
        if meerdere:
            try:
                aantal = int(count_entry.get_text().strip())
            except ValueError:
                aantal = 3
            if aantal < 1:
                aantal = 1
            if aantal > 99:
                aantal = 99
        else:
            aantal = 1

        dialog.destroy()

        if response != Gtk.ResponseType.OK:
            return

        # Genereer en print barcodes
        storage = get_backend(self.config)
        producten = storage.load_products()
        printer_cfg = self.config.get("printer", {})

        if not printer_cfg.get("enabled"):
            self._toon_foutdialoog("Printen is niet ingeschakeld in de instellingen.")
            return

        success_count = 0
        error_count = 0

        for artikelnummer in artikelnummers:
            product = next((p for p in producten if p.get("artikelnummer") == artikelnummer), None)
            if not product:
                error_count += 1
                continue

            folder = product.get("folder_locatie", "").strip()
            if not folder or not os.path.isdir(folder):
                # Maak een tijdelijke map als de productmap niet bestaat
                folder = os.path.join(self.config.get("base_folder", ""), artikelnummer)
                os.makedirs(folder, exist_ok=True)

            try:
                auto_cut = printer_cfg.get("auto_cut", True)
                for i in range(1, aantal + 1):
                    count_label = f"{i}/{aantal}" if meerdere else None
                    barcode_path = genereer_barcode(
                        artikelnummer, folder, self.config.get("barcode", {}), count_label
                    )

                    # Alleen auto-cut op de laatste label van elk product
                    if i < aantal:
                        print_barcode(barcode_path, printer_cfg, False)
                    else:
                        print_barcode(barcode_path, printer_cfg, auto_cut)
                success_count += 1
            except Exception as e:
                error_count += 1
                self._toon_foutdialoog(f"Fout bij printen {artikelnummer}: {e}")

        if success_count > 0:
            msg = f"{success_count} product(en) succesvol geprint."
            if error_count > 0:
                msg += f" {error_count} fout(en)."
            self._toon_info_dialoog(msg)

    def _export_csv(self, widget):
        """Exporteer de huidige weergave naar CSV."""

        # Kies bestandsnaam
        dialog = Gtk.FileChooserDialog(
            title="CSV opslaan",
            transient_for=self.main_window,
            action=Gtk.FileChooserAction.SAVE
        )
        dialog.add_buttons(Gtk.STOCK_CANCEL, Gtk.ResponseType.CANCEL, Gtk.STOCK_SAVE, Gtk.ResponseType.OK)
        dialog.set_current_name(f"producten_{datetime.date.today().isoformat()}.csv")

        # Voeg CSV filter toe
        filter_csv = Gtk.FileFilter()
        filter_csv.set_name("CSV-bestanden (*.csv)")
        filter_csv.add_pattern("*.csv")
        dialog.add_filter(filter_csv)

        response = dialog.run()
        if response == Gtk.ResponseType.OK:
            bestandspad = dialog.get_filename()
            if not bestandspad.endswith('.csv'):
                bestandspad += '.csv'
            self._exporteer_naar_csv(bestandspad)
        dialog.destroy()

    def _exporteer_naar_csv(self, bestandspad):
        """Exporteer de huidige tabel naar CSV."""
        try:
            # Haal alle producten op
            storage = get_backend(self.config)
            producten = storage.load_products()

            # Filter zoals in de huidige weergave
            if self.huidige_filter == FILTER_OFFLINE:
                producten = [p for p in producten
                             if p.get("verkocht", "nee").lower() != "ja"
                             and p.get("online", "nee").lower() != "ja"]
            elif self.huidige_filter == FILTER_ONLINE:
                producten = [p for p in producten
                             if p.get("verkocht", "nee").lower() != "ja"
                             and p.get("online", "nee").lower() == "ja"]
            elif self.huidige_filter == FILTER_VERKOCHT:
                producten = [p for p in producten if p.get("verkocht", "nee").lower() == "ja"]

            if self.zoekterm:
                producten = [p for p in producten
                             if self.zoekterm in p.get("artikelnummer", "").lower()
                             or self.zoekterm in p.get("titel", "").lower()]

            # Bepaal welke kolommen geëxporteerd moeten worden
            kolommen = self.user_settings.get_zichtbare_kolommen(self.huidige_filter)

            with open(bestandspad, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile, delimiter=';')

                # Schrijf headers
                headers = [KOLOM_LABELS.get(k, k) for k in kolommen]
                writer.writerow(headers)

                # Schrijf data
                for p in producten:
                    row = [p.get(k, "") for k in kolommen]
                    writer.writerow(row)

            self._toon_info_dialoog(f"CSV geëxporteerd naar:\n{bestandspad}")

        except Exception as e:
            self._toon_foutdialoog(f"Fout bij exporteren: {e}")

    def _export_excel(self, widget):
        """Exporteer naar Excel (gebruikt CSV als fallback als openpyxl niet beschikbaar is)."""
        try:
            import openpyxl
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment
        except ImportError:
            # Fallback naar CSV
            self._toon_foutdialoog(
                "openpyxl is niet geïnstalleerd.\n"
                "Installeer met: pip install openpyxl --break-system-packages\n"
                "CSV wordt geëxporteerd als alternatief."
            )
            self._export_csv(widget)
            return

        # Kies bestandsnaam
        dialog = Gtk.FileChooserDialog(
            title="Excel opslaan",
            transient_for=self.main_window,
            action=Gtk.FileChooserAction.SAVE
        )
        dialog.add_buttons(Gtk.STOCK_CANCEL, Gtk.ResponseType.CANCEL, Gtk.STOCK_SAVE, Gtk.ResponseType.OK)
        dialog.set_current_name(f"producten_{datetime.date.today().isoformat()}.xlsx")

        filter_excel = Gtk.FileFilter()
        filter_excel.set_name("Excel-bestanden (*.xlsx)")
        filter_excel.add_pattern("*.xlsx")
        dialog.add_filter(filter_excel)

        response = dialog.run()
        if response == Gtk.ResponseType.OK:
            bestandspad = dialog.get_filename()
            if not bestandspad.endswith('.xlsx'):
                bestandspad += '.xlsx'

            try:
                # Haal producten op
                storage = get_backend(self.config)
                producten = storage.load_products()

                # Filter zoals in de huidige weergave
                if self.huidige_filter == FILTER_OFFLINE:
                    producten = [p for p in producten
                                 if p.get("verkocht", "nee").lower() != "ja"
                                 and p.get("online", "nee").lower() != "ja"]
                elif self.huidige_filter == FILTER_ONLINE:
                    producten = [p for p in producten
                                 if p.get("verkocht", "nee").lower() != "ja"
                                 and p.get("online", "nee").lower() == "ja"]
                elif self.huidige_filter == FILTER_VERKOCHT:
                    producten = [p for p in producten if p.get("verkocht", "nee").lower() == "ja"]

                if self.zoekterm:
                    producten = [p for p in producten
                                 if self.zoekterm in p.get("artikelnummer", "").lower()
                                 or self.zoekterm in p.get("titel", "").lower()]

                kolommen = self.user_settings.get_zichtbare_kolommen(self.huidige_filter)

                # Maak Excel workbook
                wb = Workbook()
                ws = wb.active
                ws.title = "Producten"

                # Schrijf headers
                headers = [KOLOM_LABELS.get(k, k) for k in kolommen]
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')

                # Schrijf data
                for row, p in enumerate(producten, 2):
                    for col, k in enumerate(kolommen, 1):
                        ws.cell(row=row, column=col, value=p.get(k, ""))

                # Pas kolombreedte aan
                for col in range(1, len(headers) + 1):
                    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

                wb.save(bestandspad)
                self._toon_info_dialoog(f"Excel geëxporteerd naar:\n{bestandspad}")

            except Exception as e:
                self._toon_foutdialoog(f"Fout bij exporteren: {e}")

        dialog.destroy()

    def _open_dashboard(self, widget):
        """Open het dashboard venster."""
        storage = get_backend(self.config)
        producten = storage.load_products()

        dialog = DashboardDialog(self.main_window, producten)
        dialog.run()
        dialog.destroy()


# ============================================
# DASHBOARD DIALOOG
# ============================================
class DashboardDialog(Gtk.Dialog):
    def __init__(self, parent, producten):
        super().__init__(title="📈 Dashboard - Statistieken", transient_for=parent, flags=0)
        self.set_default_size(600, 500)
        self.set_position(Gtk.WindowPosition.CENTER)
        self.add_buttons(Gtk.STOCK_CLOSE, Gtk.ResponseType.CLOSE)

        box = self.get_content_area()
        box.set_spacing(10)
        box.set_border_width(15)

        # Scrollbaar gebied
        scrolled = Gtk.ScrolledWindow()
        scrolled.set_policy(Gtk.PolicyType.NEVER, Gtk.PolicyType.AUTOMATIC)
        box.pack_start(scrolled, True, True, 0)

        inner = Gtk.Box(orientation=Gtk.Orientation.VERTICAL, spacing=10)
        inner.set_border_width(5)
        scrolled.add(inner)

        # Totaal overzicht
        totaal = len(producten)
        online = len([p for p in producten if p.get("online", "nee").lower() == "ja"])
        offline = len([p for p in producten if p.get("online", "nee").lower() != "ja"
                       and p.get("verkocht", "nee").lower() != "ja"])
        verkocht = len([p for p in producten if p.get("verkocht", "nee").lower() == "ja"])

        totaal_label = Gtk.Label()
        totaal_label.set_markup(
            f"<b><big>📊 Overzicht</big></b>\n\n"
            f"<b>Totaal producten:</b> {totaal}\n"
            f"<b>Online:</b> {online}\n"
            f"<b>Offline (in voorraad):</b> {offline}\n"
            f"<b>Verkocht:</b> {verkocht}"
        )
        totaal_label.set_xalign(0)
        inner.pack_start(totaal_label, False, False, 0)

        inner.pack_start(Gtk.Separator(), False, False, 5)

        # Omzet statistieken
        omzet_label = Gtk.Label()
        omzet_label.set_xalign(0)
        omzet_label.set_markup("<b><big>💰 Omzet</big></b>")
        inner.pack_start(omzet_label, False, False, 0)

        totaal_omzet = 0.0
        verkocht_producten = []
        for p in producten:
            if p.get("verkocht", "nee").lower() == "ja":
                try:
                    prijs = p.get("verkoopprijs", "").replace("€", "").replace(",", ".").strip()
                    if prijs:
                        totaal_omzet += float(prijs)
                        verkocht_producten.append(p)
                except ValueError:
                    pass

        gem_prijs = totaal_omzet / len(verkocht_producten) if verkocht_producten else 0

        omzet_details = Gtk.Label()
        omzet_details.set_markup(
            f"<b>Totale omzet:</b> €{totaal_omzet:.2f}\n"
            f"<b>Aantal verkocht:</b> {len(verkocht_producten)}\n"
            f"<b>Gemiddelde verkoopprijs:</b> €{gem_prijs:.2f}"
        )
        omzet_details.set_xalign(0)
        inner.pack_start(omzet_details, False, False, 0)

        inner.pack_start(Gtk.Separator(), False, False, 5)

        # Categorie verdeling
        cat_label = Gtk.Label()
        cat_label.set_markup("<b><big>📂 Categorie verdeling</big></b>")
        cat_label.set_xalign(0)
        inner.pack_start(cat_label, False, False, 0)

        # Tel per categorie
        cat_count = {}
        for p in producten:
            cat = p.get("categorie", "Onbekend")
            cat_count[cat] = cat_count.get(cat, 0) + 1

        # Sorteer op aantal (hoogste eerst)
        sorted_cats = sorted(cat_count.items(), key=lambda x: x[1], reverse=True)
        cat_text = ""
        for cat, count in sorted_cats[:10]:  # Toon top 10
            percentage = (count / totaal * 100) if totaal > 0 else 0
            cat_text += f"<b>{cat}:</b> {count} ({percentage:.1f}%)\n"

        if not cat_text:
            cat_text = "Geen categorieën gevonden."

        cat_details = Gtk.Label()
        cat_details.set_markup(cat_text)
        cat_details.set_xalign(0)
        inner.pack_start(cat_details, False, False, 0)

        inner.pack_start(Gtk.Separator(), False, False, 5)

        # Medewerker statistieken
        user_label = Gtk.Label()
        user_label.set_markup("<b><big>👤 Medewerker statistieken</big></b>")
        user_label.set_xalign(0)
        inner.pack_start(user_label, False, False, 0)

        # Tel per medewerker
        user_count = {}
        for p in producten:
            user = p.get("verwerkt_door", "Onbekend")
            user_count[user] = user_count.get(user, 0) + 1

        user_text = ""
        for user, count in sorted(user_count.items(), key=lambda x: x[1], reverse=True):
            user_text += f"<b>{user}:</b> {count} producten\n"

        if not user_text:
            user_text = "Geen medewerkers gevonden."

        user_details = Gtk.Label()
        user_details.set_markup(user_text)
        user_details.set_xalign(0)
        inner.pack_start(user_details, False, False, 0)

        inner.pack_start(Gtk.Separator(), False, False, 5)

        # Tijdstip statistieken (aanmaak per dag)
        time_label = Gtk.Label()
        time_label.set_markup("<b><big>📅 Aanmaak per dag (laatste 7 dagen)</big></b>")
        time_label.set_xalign(0)
        inner.pack_start(time_label, False, False, 0)

        from collections import defaultdict
        day_count = defaultdict(int)
        today = datetime.date.today()

        for p in producten:
            datum_str = p.get("aanmaakdatum", "")
            try:
                datum = datetime.date.fromisoformat(datum_str)
                days_ago = (today - datum).days
                if 0 <= days_ago < 7:
                    day_count[datum] += 1
            except (ValueError, TypeError):
                pass

        time_text = ""
        for i in range(6, -1, -1):
            datum = today - datetime.timedelta(days=i)
            count = day_count.get(datum, 0)
            dag_naam = ["Ma", "Di", "Wo", "Do", "Vr", "Za", "Zo"][datum.weekday()]
            time_text += f"<b>{dag_naam} {datum.day}/{datum.month}:</b> {count} producten\n"

        if not time_text:
            time_text = "Geen aanmaakgegevens gevonden."

        time_details = Gtk.Label()
        time_details.set_markup(time_text)
        time_details.set_xalign(0)
        inner.pack_start(time_details, False, False, 0)

        self.show_all()


# ============================================
# ANDERE DIALOGEN (KolomKiezer, Online, Toewijzen, Verkocht, Bewerk)
# ============================================
class KolomKiezerDialog(Gtk.Dialog):
    def __init__(self, parent, user_settings, filter_type):
        super().__init__(title=f"Kies zichtbare kolommen - {FILTER_NAMEN[filter_type]}", transient_for=parent, flags=0)
        self.user_settings = user_settings
        self.filter_type = filter_type
        self.set_default_size(400, 450)
        self.add_buttons(Gtk.STOCK_CANCEL, Gtk.ResponseType.CANCEL, Gtk.STOCK_SAVE, Gtk.ResponseType.OK)

        box = self.get_content_area()
        box.set_spacing(8)
        box.set_border_width(12)

        box.pack_start(Gtk.Label(label=f"Selecteer welke kolommen zichtbaar moeten zijn voor filter '{FILTER_NAMEN[filter_type]}':", xalign=0), False, False, 0)
        box.pack_start(Gtk.Label(label=f"<small>Instellingen voor gebruiker: <b>{parent.gebruikersnaam}</b></small>", xalign=0), False, False, 0)

        scrolled = Gtk.ScrolledWindow()
        scrolled.set_size_request(-1, 350)
        box.pack_start(scrolled, True, True, 0)

        self.kolom_checkboxes = {}
        kolom_grid = Gtk.Grid()
        kolom_grid.set_column_spacing(20)
        kolom_grid.set_row_spacing(3)

        zichtbare = set(user_settings.get_zichtbare_kolommen(filter_type))

        sorted_kolommen = sorted(COLUMNS, key=lambda k: KOLOM_LABELS.get(k, k))

        row = 0
        col = 0
        for kolom in sorted_kolommen:
            if kolom in ["omschrijving", "algemene_voorwaarden", "folder_locatie"]:
                continue
            label = KOLOM_LABELS.get(kolom, kolom)
            check = Gtk.CheckButton(label=label)
            check.set_active(kolom in zichtbare)
            self.kolom_checkboxes[kolom] = check
            kolom_grid.attach(check, col, row, 1, 1)
            col += 1
            if col >= 3:
                col = 0
                row += 1

        scrolled.add(kolom_grid)

        btn_row = Gtk.Box(spacing=5)
        select_all_btn = Gtk.Button(label="Alles selecteren")
        select_all_btn.connect("clicked", self._on_select_all)
        btn_row.pack_start(select_all_btn, False, False, 0)

        select_none_btn = Gtk.Button(label="Niets selecteren")
        select_none_btn.connect("clicked", self._on_select_none)
        btn_row.pack_start(select_none_btn, False, False, 0)

        box.pack_start(btn_row, False, False, 5)

        self.show_all()

    def _on_select_all(self, widget):
        for check in self.kolom_checkboxes.values():
            check.set_active(True)

    def _on_select_none(self, widget):
        for check in self.kolom_checkboxes.values():
            check.set_active(False)

    def apply_to_settings(self):
        zichtbaar = [k for k, check in self.kolom_checkboxes.items() if check.get_active()]
        self.user_settings.set_zichtbare_kolommen(self.filter_type, zichtbaar)


class OnlineDialog(Gtk.Dialog):
    def __init__(self, parent, artikelnummers):
        if len(artikelnummers) == 1:
            title = f"Advertentie-URLs - {artikelnummers[0]}"
        else:
            title = f"Advertentie-URLs - {len(artikelnummers)} producten"

        super().__init__(title=title, transient_for=parent, flags=0)
        self.set_default_size(500, 350)
        self.add_buttons(Gtk.STOCK_CANCEL, Gtk.ResponseType.CANCEL, Gtk.STOCK_OK, Gtk.ResponseType.OK)

        box = self.get_content_area()
        box.set_spacing(8)
        box.set_border_width(12)

        if len(artikelnummers) == 1:
            msg = f"Voer advertentie-URL(s) in voor {artikelnummers[0]} (max 5):"
        else:
            msg = f"Voer advertentie-URL(s) in voor {len(artikelnummers)} producten (max 5):"

        box.pack_start(Gtk.Label(label=msg, xalign=0), False, False, 0)

        self.url_entries = []
        for i in range(5):
            entry = Gtk.Entry()
            entry.set_placeholder_text(f"URL {i+1} (optioneel)")
            box.pack_start(entry, False, False, 3)
            self.url_entries.append(entry)

        self.show_all()

    def get_urls(self):
        urls = []
        for entry in self.url_entries:
            url = entry.get_text().strip()
            if url:
                if not url.startswith(('http://', 'https://')):
                    url = f"https://{url}"
                urls.append(url)
        return urls


class ToewijzenDialog(Gtk.Dialog):
    def __init__(self, parent, product, user_manager, huidige_gebruiker):
        super().__init__(title=f"Toewijzen - {product.get('artikelnummer', '')}", transient_for=parent, flags=0)
        self.set_default_size(320, 150)
        self.add_buttons(Gtk.STOCK_CANCEL, Gtk.ResponseType.CANCEL, Gtk.STOCK_OK, Gtk.ResponseType.OK)

        box = self.get_content_area()
        box.set_spacing(8)
        box.set_border_width(12)

        box.pack_start(Gtk.Label(label="Selecteer de medewerker die dit product afhandelt:", xalign=0), False, False, 0)

        self.gebruikers_combo = Gtk.ComboBoxText()
        self.gebruikers_combo.append_text("")  # leeg = niet toegewezen
        for naam in user_manager.gebruikersnamen():
            self.gebruikers_combo.append_text(naam)

        huidige = product.get("toegewezen_aan", "")
        if huidige:
            try:
                self.gebruikers_combo.set_active(1 + user_manager.gebruikersnamen().index(huidige))
            except ValueError:
                self.gebruikers_combo.set_active(0)
        else:
            self.gebruikers_combo.set_active(0)

        box.pack_start(self.gebruikers_combo, False, False, 0)

        self.show_all()

    def get_toegewezen_aan(self):
        tekst = self.gebruikers_combo.get_active_text()
        return tekst.strip() if tekst else ""


class VerkochtDialog(Gtk.Dialog):
    def __init__(self, parent, artikelnummer):
        super().__init__(title=f"Markeer als verkocht - {artikelnummer}", transient_for=parent, flags=0)
        self.set_default_size(420, 380)
        self.add_buttons(Gtk.STOCK_CANCEL, Gtk.ResponseType.CANCEL, Gtk.STOCK_OK, Gtk.ResponseType.OK)

        box = self.get_content_area()
        box.set_spacing(8)
        box.set_border_width(12)

        box.pack_start(Gtk.Label(label=f"Verkoopprijs voor {artikelnummer}:", xalign=0), False, False, 0)
        self.prijs_entry = Gtk.Entry()
        self.prijs_entry.set_placeholder_text("bijv. 45.00")
        box.pack_start(self.prijs_entry, False, False, 0)

        vandaag = datetime.date.today().isoformat()
        box.pack_start(Gtk.Label(label=f"Verkoopdatum: {vandaag} (automatisch)", xalign=0), False, False, 0)
        self.vandaag = vandaag

        box.pack_start(Gtk.Separator(), False, False, 6)

        box.pack_start(self._bold_label("Leverwijze"), False, False, 0)
        lever_row = Gtk.Box(spacing=10)
        self.ophalen_btn = Gtk.RadioButton.new_with_label_from_widget(None, "Ophalen")
        self.verzenden_btn = Gtk.RadioButton.new_with_label_from_widget(self.ophalen_btn, "Verzenden")
        self.ophalen_btn.connect("toggled", self._on_leverwijze_changed)
        lever_row.pack_start(self.ophalen_btn, False, False, 0)
        lever_row.pack_start(self.verzenden_btn, False, False, 0)
        box.pack_start(lever_row, False, False, 0)

        box.pack_start(Gtk.Label(label="Marktplaatsnaam koper:", xalign=0), False, False, 0)
        self.klant_naam_entry = Gtk.Entry()
        box.pack_start(self.klant_naam_entry, False, False, 0)

        self.ophalen_box = Gtk.Box(orientation=Gtk.Orientation.VERTICAL, spacing=4)
        box.pack_start(self.ophalen_box, False, False, 0)

        self.ophalen_box.pack_start(Gtk.Label(label="Telefoonnummer:", xalign=0), False, False, 0)
        self.telefoon_entry = Gtk.Entry()
        self.ophalen_box.pack_start(self.telefoon_entry, False, False, 0)

        self.ophalen_box.pack_start(Gtk.Label(label="E-mail:", xalign=0), False, False, 0)
        self.email_entry = Gtk.Entry()
        self.ophalen_box.pack_start(self.email_entry, False, False, 0)

        self.ophalen_box.pack_start(Gtk.Label(label="Afspraak (datum/tijd ophalen):", xalign=0), False, False, 0)
        self.afspraak_entry = Gtk.Entry()
        self.afspraak_entry.set_placeholder_text("bijv. zaterdag 14:00")
        self.ophalen_box.pack_start(self.afspraak_entry, False, False, 0)

        self.verzenden_box = Gtk.Box(orientation=Gtk.Orientation.VERTICAL, spacing=4)
        box.pack_start(self.verzenden_box, False, False, 0)

        self.verzenden_box.pack_start(Gtk.Label(label="Track & Trace-nummer (optioneel):", xalign=0), False, False, 0)
        self.track_trace_entry = Gtk.Entry()
        self.verzenden_box.pack_start(self.track_trace_entry, False, False, 0)

        self.ophalen_btn.set_active(True)
        self._on_leverwijze_changed(self.ophalen_btn)

        self.show_all()

    def _bold_label(self, text):
        label = Gtk.Label()
        label.set_markup(f"<b>{text}</b>")
        label.set_xalign(0)
        return label

    def _on_leverwijze_changed(self, widget):
        is_ophalen = self.ophalen_btn.get_active()
        self.ophalen_box.set_visible(is_ophalen)
        self.verzenden_box.set_visible(not is_ophalen)

    def get_gegevens(self):
        leverwijze = "ophalen" if self.ophalen_btn.get_active() else "verzenden"
        gegevens = {
            "verkoopprijs": self.prijs_entry.get_text().strip(),
            "verkoopdatum": self.vandaag,
            "leverwijze": leverwijze,
            "klant_naam": self.klant_naam_entry.get_text().strip(),
        }
        if leverwijze == "ophalen":
            gegevens["klant_telefoon"] = self.telefoon_entry.get_text().strip()
            gegevens["klant_email"] = self.email_entry.get_text().strip()
            gegevens["ophaal_afspraak"] = self.afspraak_entry.get_text().strip()
            gegevens["track_trace"] = ""
        else:
            gegevens["klant_telefoon"] = ""
            gegevens["klant_email"] = ""
            gegevens["ophaal_afspraak"] = ""
            gegevens["track_trace"] = self.track_trace_entry.get_text().strip()
        return gegevens


class BewerkDialog(Gtk.Dialog):
    def __init__(self, parent, product, config):
        super().__init__(title=f"Bewerken - {product.get('artikelnummer', '')}", transient_for=parent, flags=0)
        self.product = dict(product)
        self.config = config
        self.set_default_size(500, 600)
        self.add_buttons(Gtk.STOCK_CANCEL, Gtk.ResponseType.CANCEL, Gtk.STOCK_SAVE, Gtk.ResponseType.OK)

        box = self.get_content_area()
        box.set_spacing(8)
        box.set_border_width(10)

        scrolled = Gtk.ScrolledWindow()
        scrolled.set_policy(Gtk.PolicyType.NEVER, Gtk.PolicyType.AUTOMATIC)
        box.pack_start(scrolled, True, True, 0)

        inner = Gtk.Box(orientation=Gtk.Orientation.VERTICAL, spacing=6)
        inner.set_border_width(5)
        scrolled.add(inner)

        self.entries = {}
        inner.pack_start(Gtk.Label(label=f"Artikelnummer: {self.product.get('artikelnummer', '')}"), False, False, 0)

        bewerkbare_velden = [c for c in COLUMNS if c not in ("artikelnummer", "folder_locatie")]
        for veld in bewerkbare_velden:
            label = Gtk.Label(label=KOLOM_LABELS.get(veld, veld))
            label.set_xalign(0)
            inner.pack_start(label, False, False, 0)

            if veld in ("omschrijving", "algemene_voorwaarden"):
                view = Gtk.TextView()
                view.get_buffer().set_text(self.product.get(veld, ""))
                view_scroll = Gtk.ScrolledWindow()
                view_scroll.set_size_request(-1, 80)
                view_scroll.add(view)
                inner.pack_start(view_scroll, False, False, 0)
                self.entries[veld] = view
            elif veld.startswith("url_"):
                entry = Gtk.Entry()
                entry.set_text(self.product.get(veld, ""))
                entry.set_placeholder_text(f"https://www.marktplaats.nl/v/...")
                inner.pack_start(entry, False, False, 0)
                self.entries[veld] = entry
            else:
                entry = Gtk.Entry()
                entry.set_text(self.product.get(veld, ""))
                inner.pack_start(entry, False, False, 0)
                self.entries[veld] = entry

        self.show_all()

    def get_product(self):
        result = dict(self.product)
        for veld, widget in self.entries.items():
            if isinstance(widget, Gtk.TextView):
                buf = widget.get_buffer()
                result[veld] = buf.get_text(buf.get_start_iter(), buf.get_end_iter(), True)
            else:
                result[veld] = widget.get_text()
        return result


# ============================================
# HOOFDVENSTER
# ============================================
class MainWindow(Gtk.Window):
    def __init__(self, gebruikersnaam, user_manager):
        super().__init__(title="Marktplaats Product Manager")
        self.set_default_size(900, 800)
        self.set_position(Gtk.WindowPosition.CENTER)

        icon_path = os.path.join(config_dir(), "icon.png")
        if os.path.exists(icon_path):
            try:
                self.set_icon_from_file(icon_path)
            except Exception:
                pass

        self.config = ConfigManager()
        self.gebruikersnaam = gebruikersnaam
        self.user_manager = user_manager
        self.user_settings = UserSettings(gebruikersnaam)

        vbox = Gtk.Box(orientation=Gtk.Orientation.VERTICAL)
        self.add(vbox)

        menu_row = Gtk.Box(spacing=5)
        menu_row.set_border_width(5)
        settings_btn = Gtk.Button(label="⚙️ Instellingen")
        settings_btn.connect("clicked", self._open_settings)
        menu_row.pack_end(settings_btn, False, False, 0)

        users_btn = Gtk.Button(label="👤 Gebruikers")
        users_btn.connect("clicked", self._open_users)
        menu_row.pack_end(users_btn, False, False, 0)

        logout_btn = Gtk.Button(label="🚪 Uitloggen / Wisselen")
        logout_btn.connect("clicked", self._on_logout)
        menu_row.pack_end(logout_btn, False, False, 0)

        gebruiker_label = Gtk.Label()
        gebruiker_label.set_markup(f"<small>Ingelogd als: <b>{GLib.markup_escape_text(gebruikersnaam)}</b></small>")
        menu_row.pack_end(gebruiker_label, False, False, 10)

        backend_label = Gtk.Label()
        backend_label.set_markup(f"<small>Actieve opslag: <b>{self.config.get('storage_backend')}</b></small>")
        self.backend_label = backend_label
        menu_row.pack_start(backend_label, False, False, 5)
        vbox.pack_start(menu_row, False, False, 0)

        self.notebook = Gtk.Notebook()
        vbox.pack_start(self.notebook, True, True, 0)

        self.registreer_tab = RegistreerTab(self)
        self.notebook.append_page(self.registreer_tab, Gtk.Label(label="📝 Registreren"))

        self.overzicht_tab = OverzichtTab(self)
        self.notebook.append_page(self.overzicht_tab, Gtk.Label(label="📊 Overzicht"))

        # Statusbalk
        self.statusbar = Gtk.Statusbar()
        # set_has_resize_grip is deprecated/verwijderd in nieuwere GTK
        # Gebruik in plaats daarvan:
        self.statusbar.set_hexpand(True)
        vbox.pack_start(self.statusbar, False, False, 0)
        self.status_context = self.statusbar.get_context_id("main")
        self.statusbar.push(self.status_context, "✅ Klaar voor gebruik")

    def _open_settings(self, widget):
        dialog = SettingsDialog(self, self.config)
        response = dialog.run()
        if response == Gtk.ResponseType.OK:
            dialog.apply_to_config()
            self.backend_label.set_markup(f"<small>Actieve opslag: <b>{self.config.get('storage_backend')}</b></small>")
            self._herbouw_registreer_tab()
            self.overzicht_tab._update_tree_columns()
            self.overzicht_tab.herlaad()
        dialog.destroy()

    def _open_users(self, widget):
        dialog = UsersDialog(self, self.user_manager, self.gebruikersnaam)
        dialog.run()
        dialog.destroy()

    def _on_logout(self, widget):
        """Uitloggen en terug naar het inlogscherm."""
        confirm = Gtk.MessageDialog(
            transient_for=self, flags=0,
            message_type=Gtk.MessageType.QUESTION, buttons=Gtk.ButtonsType.YES_NO,
            text="Weet je zeker dat je wilt uitloggen?"
        )
        confirm.format_secondary_text("Je wordt teruggebracht naar het inlogscherm.")
        response = confirm.run()
        confirm.destroy()

        if response == Gtk.ResponseType.YES:
            self.destroy()
            main()

    def _herbouw_registreer_tab(self):
        self.notebook.remove_page(0)
        self.registreer_tab = RegistreerTab(self)
        self.notebook.insert_page(self.registreer_tab, Gtk.Label(label="📝 Registreren"), 0)
        self.notebook.show_all()


def main():
    GLib.set_prgname("marktplaats_productmanager")
    apply_css()

    user_manager = UserManager()

    # Als er nog geen gebruikers zijn, maak dan een admin aan
    if not user_manager.heeft_gebruikers():
        dialog = NieuweGebruikerDialog(None, titel="Eerste gebruiker aanmaken")
        while True:
            response = dialog.run()
            if response != Gtk.ResponseType.OK:
                dialog.destroy()
                return
            naam, ww1, ww2 = dialog.get_invoer()
            if not naam:
                dialog.toon_fout("Vul een gebruikersnaam in.")
                continue
            if not ww1 or ww1 != ww2:
                dialog.toon_fout("Wachtwoorden komen niet overeen.")
                continue
            user_manager.voeg_toe(naam, ww1)
            gebruikersnaam = naam
            break
        dialog.destroy()

    login = LoginDialog(user_manager)
    gebruikersnaam = None
    while True:
        response = login.run()
        if response != Gtk.ResponseType.OK:
            login.destroy()
            return
        naam = login.get_naam()
        wachtwoord = login.get_wachtwoord()
        if not naam:
            login.toon_fout("Vul een gebruikersnaam in.")
            continue
        if user_manager.controleer(naam, wachtwoord):
            gebruikersnaam = naam
            break
        login.toon_fout("Onjuiste gebruikersnaam of wachtwoord.")
    login.destroy()

    win = MainWindow(gebruikersnaam, user_manager)
    win.connect("destroy", Gtk.main_quit)
    win.show_all()
    Gtk.main()


if __name__ == "__main__":
    main()
